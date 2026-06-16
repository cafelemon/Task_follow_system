#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import os
import re
import sys
from collections import Counter, defaultdict
from dataclasses import asdict, dataclass
from datetime import date, datetime, time, timezone
from pathlib import Path
from typing import Any


ROOT = Path(__file__).resolve().parents[1]
BACKEND = ROOT / "backend"
sys.path.insert(0, str(BACKEND))

try:
    from openpyxl import load_workbook
except ModuleNotFoundError as exc:  # pragma: no cover - command-line guard
    raise SystemExit("openpyxl 未安装，请先执行：python3 -m pip install openpyxl") from exc

from sqlalchemy import select
from sqlalchemy.orm import Session, selectinload


def load_env_file(path: Path) -> None:
    if not path.exists():
        raise FileNotFoundError(path)
    for raw_line in path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        os.environ.setdefault(key.strip(), value.strip().strip('"').strip("'"))


def bootstrap_environment() -> None:
    parser = argparse.ArgumentParser(add_help=False)
    parser.add_argument("--env-file", type=Path)
    parser.add_argument("--database-url")
    args, _ = parser.parse_known_args()
    if args.env_file:
        load_env_file(args.env_file)
    if args.database_url:
        os.environ["TASK_FOLLOW_DATABASE_URL"] = args.database_url


bootstrap_environment()

from app.db.session import SessionLocal
from app.models.entities import SubTask, TaskEvent, User, WeeklyUpdate


DEFAULT_EXCEL_PATH = ROOT / "2026公司工作任务跟踪表 (1).xlsx"
BLANK_VALUES = {"", "/", "无", "暂无", "无。", "暂无。", "无遗留事项", "无遗留事项。"}


@dataclass
class RowResult:
    status: str
    source_row: int
    task_code: str
    source_sub_code: str
    title: str
    assignee: str
    message: str
    sub_task_code: str | None = None
    weekly_update_id: int | None = None


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip()


def meaningful_text(value: Any) -> str:
    text = clean_text(value)
    return "" if text in BLANK_VALUES else text


def normalize_text(value: Any) -> str:
    return re.sub(r"\s+", " ", clean_text(value))


def split_people(value: Any) -> list[str]:
    raw = clean_text(value)
    items = [item.strip() for item in re.split(r"[\n,，、;；/]+", raw) if item.strip()]
    return list(dict.fromkeys(items))


def task_code_from_label(value: Any) -> str:
    match = re.match(r"(T-\d{3}-\d{2})\b", clean_text(value))
    return match.group(1) if match else ""


def parse_week_key(value: Any) -> str:
    raw = clean_text(value)
    match = re.search(r"W(\d{1,2})", raw, re.IGNORECASE)
    if match:
        return f"2026-W{int(match.group(1)):02d}"
    return ""


def parse_submitted_at(value: Any) -> datetime:
    if isinstance(value, datetime):
        return value if value.tzinfo else value.replace(tzinfo=timezone.utc)
    if isinstance(value, date):
        return datetime.combine(value, time(12, 0), tzinfo=timezone.utc)
    raw = clean_text(value)
    match = re.search(r"(20\d{2})[-/](\d{1,2})[-/](\d{1,2})", raw)
    if match:
        return datetime(
            int(match.group(1)),
            int(match.group(2)),
            int(match.group(3)),
            12,
            0,
            tzinfo=timezone.utc,
        )
    return datetime.now(timezone.utc)


def worksheet_rows(path: Path, sheet_name: str) -> list[tuple[int, dict[str, Any]]]:
    workbook = load_workbook(path, data_only=True)
    try:
        sheet = workbook[sheet_name]
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [clean_text(value) for value in rows[0]]
        result: list[tuple[int, dict[str, Any]]] = []
        for index, row in enumerate(rows[1:], start=2):
            if any(clean_text(value) for value in row):
                result.append((index, dict(zip(headers, row))))
        return result
    finally:
        workbook.close()


def build_source_title_map(sub_task_rows: list[tuple[int, dict[str, Any]]]) -> dict[tuple[str, str], set[str]]:
    titles: dict[tuple[str, str], set[str]] = defaultdict(set)
    for _, row in sub_task_rows:
        task_code = clean_text(row.get("任务编号"))
        source_sub_code = clean_text(row.get("子任务编号"))
        title = normalize_text(row.get("具体任务"))
        if task_code and source_sub_code and title:
            titles[(task_code, source_sub_code)].add(title)
    return titles


def existing_sub_tasks(db: Session) -> dict[tuple[str, str], list[SubTask]]:
    result: dict[tuple[str, str], list[SubTask]] = defaultdict(list)
    tasks = db.scalars(
        select(SubTask)
        .options(
            selectinload(SubTask.department_task),
            selectinload(SubTask.executors),
            selectinload(SubTask.executor),
        )
        .order_by(SubTask.code)
    ).all()
    for task in tasks:
        if task.department_task:
            result[(task.department_task.code, normalize_text(task.title))].append(task)
    return result


def resolve_sub_task(
    *,
    row: dict[str, Any],
    task_code: str,
    source_sub_code: str,
    source_titles: dict[tuple[str, str], set[str]],
    sub_tasks_by_title: dict[tuple[str, str], list[SubTask]],
    assignee_names: list[str],
) -> tuple[SubTask | None, str]:
    candidate_titles = {normalize_text(row.get("具体任务"))}
    candidate_titles.update(source_titles.get((task_code, source_sub_code), set()))
    candidate_titles.discard("")

    candidates: list[SubTask] = []
    for title in candidate_titles:
        candidates.extend(sub_tasks_by_title.get((task_code, title), []))
    unique_candidates = {item.id: item for item in candidates}
    candidates = list(unique_candidates.values())
    if not candidates:
        return None, "未匹配到系统子任务"
    if len(candidates) == 1:
        return candidates[0], ""

    assignee_set = set(assignee_names)
    matched = [
        task
        for task in candidates
        if assignee_set
        & {user.name for user in (list(task.executors or []) or [task.executor]) if user}
    ]
    if len(matched) == 1:
        return matched[0], ""
    return None, "系统内存在多条候选子任务，无法安全判断"


def resolve_assignees(sub_task: SubTask, names: list[str]) -> tuple[list[User], list[str]]:
    existing = {user.name: user for user in (list(sub_task.executors or []) or [sub_task.executor]) if user}
    matched: list[User] = []
    missing: list[str] = []
    for name in names:
        user = existing.get(name)
        if user:
            matched.append(user)
        else:
            missing.append(name)
    return matched, missing


def run_backfill(
    db: Session,
    *,
    excel_path: Path,
    week_key: str,
    apply: bool,
) -> dict[str, Any]:
    sub_task_rows = worksheet_rows(excel_path, "03_部门拆解任务")
    weekly_rows = worksheet_rows(excel_path, "04_周更新进度")
    source_titles = build_source_title_map(sub_task_rows)
    sub_tasks_by_title = existing_sub_tasks(db)

    results: list[RowResult] = []
    inserted = 0

    for source_row, row in weekly_rows:
        if parse_week_key(row.get("所属周")) != week_key:
            continue
        task_code = task_code_from_label(row.get("任务项"))
        source_sub_code = clean_text(row.get("子任务编号"))
        title = normalize_text(row.get("具体任务"))
        this_week = meaningful_text(row.get("本周完成内容"))
        next_week = meaningful_text(row.get("下周工作计划"))
        risk = meaningful_text(row.get("遗留事项"))
        if not any([this_week, next_week, risk]):
            results.append(RowResult("skipped_empty", source_row, task_code, source_sub_code, title, "", "本周更新内容为空"))
            continue
        if not task_code or not source_sub_code:
            results.append(RowResult("unresolved_task", source_row, task_code, source_sub_code, title, "", "缺少部门任务编号或原始子任务编号"))
            continue

        assignee_names = split_people(row.get("执行责任人"))
        sub_task, task_error = resolve_sub_task(
            row=row,
            task_code=task_code,
            source_sub_code=source_sub_code,
            source_titles=source_titles,
            sub_tasks_by_title=sub_tasks_by_title,
            assignee_names=assignee_names,
        )
        if not sub_task:
            results.append(RowResult("unresolved_task", source_row, task_code, source_sub_code, title, "", task_error))
            continue

        if not assignee_names:
            results.append(
                RowResult(
                    "unresolved_assignee",
                    source_row,
                    task_code,
                    source_sub_code,
                    title,
                    "",
                    "缺少执行责任人",
                    sub_task.code,
                )
            )
            continue

        assignees, missing = resolve_assignees(sub_task, assignee_names)
        for name in missing:
            results.append(
                RowResult(
                    "unresolved_assignee",
                    source_row,
                    task_code,
                    source_sub_code,
                    title,
                    name,
                    "执行责任人与系统子任务执行人不匹配",
                    sub_task.code,
                )
            )
        for assignee in assignees:
            existing = db.scalar(
                select(WeeklyUpdate).where(
                    WeeklyUpdate.sub_task_id == sub_task.id,
                    WeeklyUpdate.week_key == week_key,
                    WeeklyUpdate.assignee_id == assignee.id,
                )
            )
            if existing:
                results.append(
                    RowResult(
                        "skipped_existing",
                        source_row,
                        task_code,
                        source_sub_code,
                        title,
                        assignee.name,
                        "系统中已存在该周更新，按只补空不覆盖策略跳过",
                        sub_task.code,
                        existing.id,
                    )
                )
                continue
            update_id: int | None = None
            if apply:
                update = WeeklyUpdate(
                    sub_task_id=sub_task.id,
                    assignee_id=assignee.id,
                    week_key=week_key,
                    status="submitted",
                    progress=sub_task.progress or 0,
                    this_week=this_week,
                    next_week=next_week,
                    risk=risk,
                    risk_level=None,
                    needs_coordination=False,
                    submitter_id=assignee.id,
                    submitted_at=parse_submitted_at(row.get("更新日期")),
                )
                db.add(update)
                db.flush()
                db.add(
                    TaskEvent(
                        object_type="sub_task",
                        object_id=sub_task.id,
                        event_type="weekly_update_backfilled",
                        title="4.8.0 回填周更新",
                        content=f"{week_key} / {assignee.name} / {excel_path.name}",
                        actor_id=None,
                    )
                )
                update_id = update.id
                inserted += 1
            results.append(
                RowResult(
                    "inserted" if apply else "ready",
                    source_row,
                    task_code,
                    source_sub_code,
                    title,
                    assignee.name,
                    "可写入" if not apply else "已写入",
                    sub_task.code,
                    update_id,
                )
            )

    if apply:
        db.commit()
    else:
        db.rollback()

    counts = Counter(item.status for item in results)
    return {
        "ok": True,
        "mode": "apply" if apply else "dry-run",
        "excel_path": str(excel_path),
        "week_key": week_key,
        "summary": {
            "source_weekly_rows": sum(1 for _, row in weekly_rows if parse_week_key(row.get("所属周")) == week_key),
            "inserted": inserted,
            "counts": dict(sorted(counts.items())),
        },
        "results": [asdict(item) for item in results],
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Backfill one week of weekly updates from exported Base Excel snapshot.")
    parser.add_argument("--path", type=Path, default=DEFAULT_EXCEL_PATH)
    parser.add_argument("--env-file", type=Path, help="可选：先读取本地 env 文件，再初始化数据库连接。")
    parser.add_argument("--database-url", help="可选：显式指定 TASK_FOLLOW_DATABASE_URL，优先级高于 env 文件。")
    parser.add_argument("--week-key", required=True, help="例如 2026-W24；本脚本不会自动推断写入周次。")
    parser.add_argument("--apply", action="store_true", help="真实写入数据库；不传则只 dry-run。")
    parser.add_argument("--output-json", type=Path, help="可选：把完整报告写入 JSON 文件。")
    args = parser.parse_args()

    if not args.path.exists():
        raise FileNotFoundError(args.path)
    if not re.fullmatch(r"20\d{2}-W\d{2}", args.week_key):
        raise SystemExit("--week-key 必须形如 2026-W24")

    with SessionLocal() as db:
        try:
            report = run_backfill(db, excel_path=args.path, week_key=args.week_key, apply=args.apply)
        except Exception:
            db.rollback()
            raise

    payload = json.dumps(report, ensure_ascii=False, indent=2)
    if args.output_json:
        args.output_json.write_text(payload + "\n", encoding="utf-8")
    print(payload)


if __name__ == "__main__":
    main()
