import csv
import logging
import re
import uuid
from collections import defaultdict
from datetime import date, datetime, timedelta, timezone
from io import BytesIO, StringIO
from pathlib import Path
from urllib.parse import quote

from fastapi import APIRouter, Cookie, Depends, File, HTTPException, Request, Response, UploadFile, status
from fastapi.responses import FileResponse, RedirectResponse
from sqlalchemy import delete, func, or_, select, update
from sqlalchemy.orm import Session
from openpyxl import load_workbook

from app import __version__
from app.core.config import settings
from app.db.session import get_db
from app.models.entities import (
    Attachment,
    CoordinationItem,
    Department,
    DepartmentTask,
    DepartmentTaskDepartment,
    DepartmentTaskOwner,
    NotificationRecord,
    ParentTask,
    ParentTaskOwner,
    Permission,
    RiskItem,
    Role,
    StrategicGoal,
    SubTask,
    SubTaskExecutor,
    SubTaskOwner,
    TaskEvent,
    User,
    UserGuideProgress,
    WeeklyUpdate,
)
from app.schemas.dto import (
    DepartmentCreate,
    DepartmentTaskCreate,
    DepartmentTaskUpdate,
    DepartmentUpdate,
    GoalCreate,
    GuideProgressUpdate,
    LoginRequest,
    OnboardingUpdate,
    OpenIdLoginRequest,
    ParentTaskCreate,
    ParentTaskUpdate,
    PersonCreate,
    PersonUpdate,
    RiskItemCreate,
    RiskItemUpdate,
    RolePermissionUpdate,
    SubTaskCreate,
    SubTaskUpdate,
    WeeklyReminderRequest,
    WeeklyUpdateUpsert,
)
from app.services.business import (
    build_meeting_board,
    current_week_key,
    executor_people,
    generate_code,
    owner_people,
    people_payload,
    send_department_task_due_reminders,
    send_department_task_split_notifications,
    send_risk_item_notifications,
    send_risk_overdue_reminders,
    send_weekly_update_reminders,
    serialize_department_task,
    serialize_department_task_tree,
    serialize_goal,
    serialize_parent_task,
    serialize_sub_task,
    serialize_risk_item,
    serialize_user,
    serialize_weekly_update,
    risk_score,
    upsert_weekly_update,
)
from app.services.permissions import (
    can_access_parent_task,
    can_access_department_task,
    can_edit_parent_task,
    can_create_department_task,
    can_edit_department_task,
    can_access_attachment,
    can_delete_attachment,
    can_split_sub_task,
    can_access_sub_task,
    can_reopen_sub_task,
    can_update_sub_task_weekly,
    can_upload_weekly_update_attachment,
    can_view_sub_task_execution_entry,
    can_manage_parent_tasks,
    can_view_parent_task_page,
    can_view_department_directory,
    get_current_user,
    refresh_role_permissions,
    require_admin,
    require_permission,
    sub_task_executor_ids,
    sub_task_execution_relation,
    task_owner_ids,
    user_permission_codes,
    user_role_codes,
)
from app.services.auth import (
    SESSION_COOKIE,
    create_lark_oauth_authorize_url,
    create_session,
    delete_session,
    session_cookie_kwargs,
    verify_lark_oauth_state,
    verify_lark_login_token,
    verify_password,
)
from app.services.lark_client import lark_client
from app.services.lark_sync import import_base_2026, preview_base_2026
from app.services.people_department_sync import count_open_id_candidates, sync_people_from_rows
from app.services.scheduler import scheduler_status

router = APIRouter(prefix="/api")
logger = logging.getLogger(__name__)
MAX_ATTACHMENT_BYTES = 20 * 1024 * 1024
ALLOWED_ATTACHMENT_EXTENSIONS = {
    ".csv",
    ".doc",
    ".docx",
    ".jpeg",
    ".jpg",
    ".pdf",
    ".png",
    ".ppt",
    ".pptx",
    ".txt",
    ".xls",
    ".xlsx",
    ".zip",
}
CURRENT_ONBOARDING_VERSION = "1"
EXECUTIVE_FRAMEWORK_GUIDE = ("executive_framework", "1")
EXECUTIVE_MEETING_GUIDE = ("executive_meeting_board", "1")
DEPARTMENT_OWNER_FRAMEWORK_GUIDE = ("department_owner_framework", "1")
DEPARTMENT_OWNER_PARENT_TASKS_GUIDE = ("department_owner_parent_tasks", "1")
DEPARTMENT_OWNER_DEPARTMENT_TASKS_GUIDE = ("department_owner_department_tasks", "1")
DEPARTMENT_OWNER_SUB_TASKS_GUIDE = ("department_owner_sub_tasks", "1")
TASK_OWNER_FRAMEWORK_GUIDE = ("task_owner_framework", "1")
TASK_OWNER_DEPARTMENT_TASKS_GUIDE = ("task_owner_department_tasks", "1")
TASK_OWNER_SUB_TASKS_GUIDE = ("task_owner_sub_tasks", "1")
EXECUTOR_FRAMEWORK_GUIDE = ("executor_framework", "1")
EXECUTOR_SUB_TASKS_GUIDE = ("executor_sub_tasks", "1")
OBSERVER_FRAMEWORK_GUIDE = ("observer_framework", "1")
OBSERVER_MEETING_GUIDE = ("observer_meeting_board", "1")
OBSERVER_PARENT_TASKS_GUIDE = ("observer_parent_tasks", "1")
OBSERVER_DEPARTMENT_TASKS_GUIDE = ("observer_department_tasks", "1")
OBSERVER_TIMELINE_GUIDE = ("observer_timeline", "1")
OBSERVER_SUB_TASKS_GUIDE = ("observer_sub_tasks", "1")


def feature_payload(db: Session, user: User) -> dict:
    can_manage_parent = can_manage_parent_tasks(user)
    return {
        "can_view_parent_tasks": can_view_parent_task_page(db, user),
        "can_create_parent_tasks": can_manage_parent,
        "can_delete_parent_tasks": can_manage_parent,
        "can_manage_parent_tasks": can_manage_parent,
        "can_switch_department": can_view_department_directory(user),
    }


def onboarding_payload(user: User) -> dict:
    return {
        "version": CURRENT_ONBOARDING_VERSION,
        "required": user.onboarding_version != CURRENT_ONBOARDING_VERSION,
        "status": user.onboarding_status,
        "completed_at": (
            user.onboarding_completed_at.isoformat() if user.onboarding_completed_at else None
        ),
    }


def guide_profile(db: Session, user: User) -> str | None:
    roles = user_role_codes(user)
    if roles & {"general_manager", "secretary"}:
        return "executive_office"
    if "observer" in roles:
        return "observer"
    if "department_owner" in roles and user.department_id:
        return "department_owner"
    owns_department_task = db.scalar(
        select(DepartmentTaskOwner.user_id)
        .join(DepartmentTask, DepartmentTask.id == DepartmentTaskOwner.department_task_id)
        .join(ParentTask, ParentTask.id == DepartmentTask.parent_task_id)
        .where(
            DepartmentTaskOwner.user_id == user.id,
            DepartmentTask.status != "archived",
            ParentTask.status != "archived",
        )
        .limit(1)
    )
    if "task_owner" in roles or owns_department_task:
        return "task_owner"
    executes_sub_task = db.scalar(
        select(SubTaskExecutor.user_id)
        .join(SubTask, SubTask.id == SubTaskExecutor.sub_task_id)
        .join(DepartmentTask, DepartmentTask.id == SubTask.department_task_id)
        .join(ParentTask, ParentTask.id == DepartmentTask.parent_task_id)
        .where(
            SubTaskExecutor.user_id == user.id,
            SubTask.status != "archived",
            DepartmentTask.status != "archived",
            ParentTask.status != "archived",
        )
        .limit(1)
    )
    if "executor" in roles or executes_sub_task:
        return "executor"
    return None


def has_active_execution_task(db: Session, user: User) -> bool:
    return bool(
        db.scalar(
            select(SubTaskExecutor.user_id)
            .join(SubTask, SubTask.id == SubTaskExecutor.sub_task_id)
            .join(DepartmentTask, DepartmentTask.id == SubTask.department_task_id)
            .join(ParentTask, ParentTask.id == DepartmentTask.parent_task_id)
            .where(
                SubTaskExecutor.user_id == user.id,
                SubTask.status != "archived",
                DepartmentTask.status != "archived",
                ParentTask.status != "archived",
            )
            .limit(1)
        )
    )


def has_active_owned_or_execution_sub_task(db: Session, user: User) -> bool:
    owned = db.scalar(
        select(SubTaskOwner.user_id)
        .join(SubTask, SubTask.id == SubTaskOwner.sub_task_id)
        .join(DepartmentTask, DepartmentTask.id == SubTask.department_task_id)
        .join(ParentTask, ParentTask.id == DepartmentTask.parent_task_id)
        .where(
            SubTaskOwner.user_id == user.id,
            SubTask.status != "archived",
            DepartmentTask.status != "archived",
            ParentTask.status != "archived",
        )
        .limit(1)
    )
    return bool(owned or has_active_execution_task(db, user))


def guide_state(db: Session, user: User, guide_key: str, version: str) -> dict:
    progress = db.scalar(
        select(UserGuideProgress).where(
            UserGuideProgress.user_id == user.id,
            UserGuideProgress.guide_key == guide_key,
            UserGuideProgress.version == version,
        )
    )
    return {
        "guide_key": guide_key,
        "version": version,
        "required": progress is None,
        "status": progress.status if progress else None,
        "completed_at": progress.completed_at.isoformat() if progress else None,
    }


def guides_payload(db: Session, user: User) -> dict:
    profile = guide_profile(db, user)
    modules: dict[str, dict] = {}
    if profile == "executive_office":
        framework_key, framework_version = EXECUTIVE_FRAMEWORK_GUIDE
        meeting_key, meeting_version = EXECUTIVE_MEETING_GUIDE
        modules["meeting_board"] = guide_state(db, user, meeting_key, meeting_version)
        return {
            "profile": profile,
            "system": guide_state(db, user, framework_key, framework_version),
            "modules": modules,
        }
    if profile == "department_owner":
        framework_key, framework_version = DEPARTMENT_OWNER_FRAMEWORK_GUIDE
        parent_key, parent_version = DEPARTMENT_OWNER_PARENT_TASKS_GUIDE
        department_key, department_version = DEPARTMENT_OWNER_DEPARTMENT_TASKS_GUIDE
        modules["parent_tasks"] = guide_state(db, user, parent_key, parent_version)
        modules["department_tasks"] = guide_state(db, user, department_key, department_version)
        if has_active_execution_task(db, user):
            sub_key, sub_version = DEPARTMENT_OWNER_SUB_TASKS_GUIDE
            modules["sub_tasks"] = guide_state(db, user, sub_key, sub_version)
        return {
            "profile": profile,
            "system": guide_state(db, user, framework_key, framework_version),
            "modules": modules,
        }
    if profile == "task_owner":
        framework_key, framework_version = TASK_OWNER_FRAMEWORK_GUIDE
        department_key, department_version = TASK_OWNER_DEPARTMENT_TASKS_GUIDE
        modules["department_tasks"] = guide_state(db, user, department_key, department_version)
        if has_active_owned_or_execution_sub_task(db, user):
            sub_key, sub_version = TASK_OWNER_SUB_TASKS_GUIDE
            modules["sub_tasks"] = guide_state(db, user, sub_key, sub_version)
        return {
            "profile": profile,
            "system": guide_state(db, user, framework_key, framework_version),
            "modules": modules,
        }
    if profile == "executor":
        framework_key, framework_version = EXECUTOR_FRAMEWORK_GUIDE
        if has_active_execution_task(db, user):
            sub_key, sub_version = EXECUTOR_SUB_TASKS_GUIDE
            modules["sub_tasks"] = guide_state(db, user, sub_key, sub_version)
        return {
            "profile": profile,
            "system": guide_state(db, user, framework_key, framework_version),
            "modules": modules,
        }
    if profile == "observer":
        framework_key, framework_version = OBSERVER_FRAMEWORK_GUIDE
        meeting_key, meeting_version = OBSERVER_MEETING_GUIDE
        parent_key, parent_version = OBSERVER_PARENT_TASKS_GUIDE
        department_key, department_version = OBSERVER_DEPARTMENT_TASKS_GUIDE
        timeline_key, timeline_version = OBSERVER_TIMELINE_GUIDE
        modules["meeting_board"] = guide_state(db, user, meeting_key, meeting_version)
        modules["parent_tasks"] = guide_state(db, user, parent_key, parent_version)
        modules["department_tasks"] = guide_state(db, user, department_key, department_version)
        modules["timeline"] = guide_state(db, user, timeline_key, timeline_version)
        if has_active_execution_task(db, user):
            sub_key, sub_version = OBSERVER_SUB_TASKS_GUIDE
            modules["sub_tasks"] = guide_state(db, user, sub_key, sub_version)
        return {
            "profile": profile,
            "system": guide_state(db, user, framework_key, framework_version),
            "modules": modules,
        }
    return {"profile": profile, "system": None, "modules": {}}


def allowed_guides(db: Session, user: User) -> set[tuple[str, str]]:
    profile = guide_profile(db, user)
    if profile == "executive_office":
        return {EXECUTIVE_FRAMEWORK_GUIDE, EXECUTIVE_MEETING_GUIDE}
    if profile == "department_owner":
        allowed = {
            DEPARTMENT_OWNER_FRAMEWORK_GUIDE,
            DEPARTMENT_OWNER_PARENT_TASKS_GUIDE,
            DEPARTMENT_OWNER_DEPARTMENT_TASKS_GUIDE,
        }
        if has_active_execution_task(db, user):
            allowed.add(DEPARTMENT_OWNER_SUB_TASKS_GUIDE)
        return allowed
    if profile == "task_owner":
        allowed = {
            TASK_OWNER_FRAMEWORK_GUIDE,
            TASK_OWNER_DEPARTMENT_TASKS_GUIDE,
        }
        if has_active_owned_or_execution_sub_task(db, user):
            allowed.add(TASK_OWNER_SUB_TASKS_GUIDE)
        return allowed
    if profile == "executor":
        allowed = {EXECUTOR_FRAMEWORK_GUIDE}
        if has_active_execution_task(db, user):
            allowed.add(EXECUTOR_SUB_TASKS_GUIDE)
        return allowed
    if profile == "observer":
        allowed = {
            OBSERVER_FRAMEWORK_GUIDE,
            OBSERVER_MEETING_GUIDE,
            OBSERVER_PARENT_TASKS_GUIDE,
            OBSERVER_DEPARTMENT_TASKS_GUIDE,
            OBSERVER_TIMELINE_GUIDE,
        }
        if has_active_execution_task(db, user):
            allowed.add(OBSERVER_SUB_TASKS_GUIDE)
        return allowed
    return set()


def normalize_open_id(value: str | None) -> str | None:
    if value is None:
        return None
    normalized = value.strip()
    return normalized or None


def normalize_email(value: str | None) -> str | None:
    if value is None:
        return None
    normalized = value.strip().lower()
    return normalized or None


def apply_open_id_binding(db: Session, user: User, open_id: str | None) -> None:
    normalized = normalize_open_id(open_id)
    if normalized:
        exists = db.scalar(select(User).where(User.open_id == normalized, User.id != user.id))
        if exists:
            raise HTTPException(status.HTTP_409_CONFLICT, detail="open_id 已绑定到其他人员")
    if normalized != user.open_id:
        user.open_id = normalized
        user.open_id_bound_at = datetime.now(timezone.utc) if normalized else None


def apply_email_binding(db: Session, user: User, email: str | None) -> None:
    normalized = normalize_email(email)
    if normalized:
        exists = db.scalar(select(User).where(User.email == normalized, User.id != user.id))
        if exists:
            raise HTTPException(status.HTTP_409_CONFLICT, detail="邮箱已绑定到其他人员")
    user.email = normalized


def resolve_users(db: Session, user_ids: list[int] | None, fallback_id: int | None, field_name: str) -> list[User]:
    selected_ids = list(dict.fromkeys(user_ids or ([] if fallback_id is None else [fallback_id])))
    if not selected_ids:
        raise HTTPException(status.HTTP_422_UNPROCESSABLE_ENTITY, detail=f"{field_name}不能为空")
    users = list(db.scalars(select(User).where(User.id.in_(selected_ids), User.status != "disabled")).all())
    found_ids = {user.id for user in users}
    if found_ids != set(selected_ids):
        raise HTTPException(status.HTTP_404_NOT_FOUND, detail="User not found")
    return sorted(users, key=lambda item: selected_ids.index(item.id))


def sync_department_task_departments(db: Session, task: DepartmentTask, departments: list[Department]) -> None:
    task.department_id = departments[0].id
    db.flush()
    db.execute(
        delete(DepartmentTaskDepartment).where(DepartmentTaskDepartment.department_task_id == task.id)
    )
    db.add_all(
        DepartmentTaskDepartment(department_task_id=task.id, department_id=department.id)
        for department in departments
    )


def sync_task_owners(db: Session, task: ParentTask | DepartmentTask | SubTask, users: list[User]) -> None:
    task.owner_id = users[0].id
    db.flush()
    if isinstance(task, ParentTask):
        db.execute(delete(ParentTaskOwner).where(ParentTaskOwner.parent_task_id == task.id))
        db.add_all(ParentTaskOwner(parent_task_id=task.id, user_id=user.id) for user in users)
    elif isinstance(task, DepartmentTask):
        db.execute(delete(DepartmentTaskOwner).where(DepartmentTaskOwner.department_task_id == task.id))
        db.add_all(DepartmentTaskOwner(department_task_id=task.id, user_id=user.id) for user in users)
    else:
        db.execute(delete(SubTaskOwner).where(SubTaskOwner.sub_task_id == task.id))
        db.add_all(SubTaskOwner(sub_task_id=task.id, user_id=user.id) for user in users)


def sync_department_sub_task_owners(db: Session, task: DepartmentTask, users: list[User]) -> None:
    for sub_task in task.sub_tasks:
        if sub_task.status != "archived":
            sync_task_owners(db, sub_task, users)


def validate_inherited_sub_task_owners(payload: SubTaskCreate | SubTaskUpdate, owners: list[User]) -> None:
    requested_ids: list[int] | None = None
    if payload.owner_ids is not None:
        requested_ids = list(dict.fromkeys(payload.owner_ids))
    elif payload.owner_id is not None:
        requested_ids = [payload.owner_id]
    if requested_ids is not None and requested_ids != [owner.id for owner in owners]:
        raise HTTPException(
            status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="子任务负责人必须与所属部门任务负责人一致",
        )


def sync_sub_task_executors(db: Session, task: SubTask, users: list[User]) -> None:
    task.executor_id = users[0].id
    db.flush()
    db.execute(delete(SubTaskExecutor).where(SubTaskExecutor.sub_task_id == task.id))
    db.add_all(SubTaskExecutor(sub_task_id=task.id, user_id=user.id) for user in users)


def expire_task_people(db: Session, task: ParentTask | DepartmentTask | SubTask, include_executors: bool = False) -> None:
    attrs = ["owner", "owners"]
    if include_executors:
        attrs.extend(["executor", "executors"])
    db.expire(task, [attr for attr in attrs if hasattr(task, attr)])


def default_update_assignee(current_user: User, sub_task: SubTask, assignee_id: int | None) -> User:
    executors = executor_people(sub_task)
    executor_ids = {user.id for user in executors}
    if assignee_id is None:
        if current_user.id in executor_ids:
            assignee_id = current_user.id
        elif executors:
            assignee_id = executors[0].id
        else:
            assignee_id = sub_task.executor_id
    assignee = next((user for user in executors if user.id == assignee_id), None)
    if not assignee:
        raise HTTPException(status.HTTP_404_NOT_FOUND, detail="Assignee not found")
    return assignee


def attachment_storage_root() -> Path:
    root = Path(settings.attachment_root)
    if not root.is_absolute():
        root = (Path(__file__).resolve().parents[3] / root).resolve()
    return root.resolve()


def sanitize_filename(filename: str) -> str:
    name = Path(filename or "attachment").name
    name = re.sub(r"[\x00-\x1f/\\]+", "_", name).strip().strip(".")
    return name[:180] or "attachment"


def serialize_attachment_for_user(attachment: Attachment, user: User) -> dict:
    return {
        "id": attachment.id,
        "filename": attachment.filename,
        "download_url": f"/api/attachments/{attachment.id}/download",
        "can_delete": can_delete_attachment(user, attachment),
        "created_at": attachment.created_at.isoformat() if attachment.created_at else None,
        "uploader_id": attachment.uploader_id,
    }


def weekly_update_attachments(db: Session, update_id: int | None) -> list[Attachment]:
    if not update_id:
        return []
    return list(
        db.scalars(
            select(Attachment)
            .where(Attachment.related_type == "weekly_update", Attachment.related_id == update_id)
            .order_by(Attachment.id)
        ).all()
    )


def is_path_inside(path: Path, root: Path) -> bool:
    try:
        path.resolve().relative_to(root.resolve())
        return True
    except ValueError:
        return False


def attachment_file_exists(attachment: Attachment) -> bool:
    root = attachment_storage_root()
    file_path = (root / attachment.storage_path).resolve()
    return is_path_inside(file_path, root) and file_path.exists() and file_path.is_file()


def serialize_weekly_update_for_user(db: Session, update: WeeklyUpdate, user: User) -> dict:
    return {
        **serialize_weekly_update(update),
        "attachments": [
            serialize_attachment_for_user(attachment, user)
            for attachment in weekly_update_attachments(db, update.id)
            if attachment_file_exists(attachment)
        ],
    }


def lark_login_error_redirect(message: str) -> RedirectResponse:
    return RedirectResponse(
        url=f"/login?lark_error={quote(message)}",
        status_code=status.HTTP_302_FOUND,
    )


def generate_sub_task_code(db: Session, department_task: DepartmentTask) -> str:
    prefix = f"{department_task.code}-"
    existing = {
        code
        for code in db.scalars(select(SubTask.code).where(SubTask.code.like(f"{prefix}%"))).all()
        if code
    }
    index = len(existing) + 1
    while True:
        code = f"{prefix}{index:02d}"
        if code not in existing and not db.scalar(select(SubTask.id).where(SubTask.code == code)):
            return code
        index += 1


def resolve_departments(db: Session, department_id: int | None, department_ids: list[int] | None) -> tuple[int, list[Department]]:
    selected_ids = list(dict.fromkeys(department_ids or ([] if department_id is None else [department_id])))
    if not selected_ids:
        raise HTTPException(status.HTTP_422_UNPROCESSABLE_ENTITY, detail="负责部门不能为空")
    departments = list(db.scalars(select(Department).where(Department.id.in_(selected_ids))).all())
    found_ids = {department.id for department in departments}
    if found_ids != set(selected_ids):
        raise HTTPException(status.HTTP_404_NOT_FOUND, detail="Department not found")
    ordered_departments = sorted(departments, key=lambda item: selected_ids.index(item.id))
    return selected_ids[0], ordered_departments


def serialize_department_task_for_user(db: Session, user: User, task: DepartmentTask) -> dict:
    return {
        **serialize_department_task(task),
        "can_edit": can_edit_department_task(db, user, task),
        "can_delete": can_edit_department_task(db, user, task),
        "can_split": can_split_sub_task(db, user, task),
    }


def serialize_parent_task_for_user(db: Session, user: User, task: ParentTask) -> dict:
    return {
        **serialize_parent_task(task),
        "can_edit": can_edit_parent_task(user, task),
        "can_create_department_task": can_create_department_task(user, task),
    }


def updates_by_sub_task(db: Session, sub_task_ids: list[int], week_key: str) -> dict[int, list[WeeklyUpdate]]:
    if not sub_task_ids:
        return {}
    result: dict[int, list[WeeklyUpdate]] = defaultdict(list)
    for update in db.scalars(
        select(WeeklyUpdate).where(WeeklyUpdate.week_key == week_key, WeeklyUpdate.sub_task_id.in_(sub_task_ids))
    ).all():
        result[update.sub_task_id].append(update)
    return result


def merge_weekly_update_field(updates: list[WeeklyUpdate], field_name: str) -> str | None:
    parts = []
    for update in sorted(updates, key=lambda item: ((item.assignee.name if item.assignee else ""), item.id)):
        value = getattr(update, field_name)
        if value:
            assignee = update.assignee.name if update.assignee else "未指定"
            parts.append(f"{assignee}：{value}")
    return "；".join(parts) if parts else None


def weekly_update_summary(updates: list[WeeklyUpdate]) -> dict:
    return {
        "weekly_this_week": merge_weekly_update_field(updates, "this_week"),
        "weekly_risk": merge_weekly_update_field(updates, "risk"),
    }


def serialize_sub_task_with_weekly_summary(
    task: SubTask,
    updates: list[WeeklyUpdate] | None = None,
    current_update: WeeklyUpdate | None = None,
) -> dict:
    update_list = updates or ([] if current_update is None else [current_update])
    return {
        **serialize_sub_task(task, current_update or (update_list[0] if update_list else None)),
        **weekly_update_summary(update_list),
    }


def can_create_risk_item(user: User, task: SubTask) -> bool:
    if user.is_admin or "permission.manage" in user_permission_codes(user):
        return True
    if user.id in sub_task_executor_ids(task):
        return True
    if user.id in task_owner_ids(task):
        return True
    department_task = task.department_task
    if department_task and user.id in task_owner_ids(department_task):
        return True
    return False


def can_manage_risk_item(user: User, item: RiskItem) -> bool:
    if user.is_admin or "permission.manage" in user_permission_codes(user):
        return True
    if item.owner_id == user.id:
        return True
    task = item.sub_task
    if task and user.id in task_owner_ids(task):
        return True
    if task and task.department_task and user.id in task_owner_ids(task.department_task):
        return True
    return False


def risk_owner_candidates(task: SubTask) -> list[User]:
    return [owner for owner in owner_people(task) if owner.status != "disabled"]


def risk_owner_options(task: SubTask) -> list[dict]:
    return [{"id": owner.id, "name": owner.name} for owner in risk_owner_candidates(task)]


def serialize_risk_item_for_user(item: RiskItem, user: User) -> dict:
    return {
        **serialize_risk_item(item),
        "can_manage": can_manage_risk_item(user, item),
        "owner_options": risk_owner_options(item.sub_task) if item.sub_task else [],
    }


def active_risk_items_for_tasks(db: Session, tasks: list[SubTask]) -> list[RiskItem]:
    task_ids = [task.id for task in tasks]
    if not task_ids:
        return []
    return list(
        db.scalars(
            select(RiskItem).where(
                RiskItem.sub_task_id.in_(task_ids),
                RiskItem.status.in_(["open", "in_progress"]),
            )
        ).all()
    )


def risk_item_detail_summary(item: RiskItem, user: User) -> dict:
    task = item.sub_task
    department_task = task.department_task if task else None
    parent_task = department_task.parent_task if department_task else None
    overdue = bool(item.due_date and item.due_date < date.today() and item.status != "closed")
    return {
        "id": item.id,
        "code": item.code,
        "title": item.title,
        "description": item.description,
        "level": item.level,
        "risk_level": item.level,
        "impact_score": item.impact_score,
        "likelihood_score": item.likelihood_score,
        "score": item.score,
        "status": item.status,
        "owner": item.owner.name if item.owner else None,
        "owner_id": item.owner_id,
        "owner_options": risk_owner_options(task) if task else [],
        "can_manage": can_manage_risk_item(user, item),
        "due_date": item.due_date.isoformat() if item.due_date else None,
        "resolution_note": item.resolution_note,
        "is_overdue": overdue,
        "issue_type": "风险逾期" if overdue else "风险",
        "sub_task_id": item.sub_task_id,
        "sub_task_code": task.code if task else None,
        "sub_task": task.title if task else None,
        "department_task": department_task.title if department_task else None,
        "department_task_code": department_task.code if department_task else None,
        "parent_task": parent_task.title if parent_task else None,
        "parent_task_code": parent_task.code if parent_task else None,
        "created_at": item.created_at.isoformat() if item.created_at else None,
    }


def serialize_department_task_tree_for_user(db: Session, user: User, task: DepartmentTask) -> dict:
    current_week = current_week_key()
    updates = updates_by_sub_task(db, [sub_task.id for sub_task in task.sub_tasks], current_week)
    return {
        **serialize_department_task_for_user(db, user, task),
        "sub_tasks": [
            serialize_sub_task_with_weekly_summary(sub_task, updates.get(sub_task.id, []))
            for sub_task in sorted(task.sub_tasks, key=lambda item: item.code)
        ],
    }


def week_key_from_date(value: date) -> str:
    year, week, _ = value.isocalendar()
    return f"{year}-W{week:02d}"


def monday_from_week_key(week_key: str) -> date:
    year_text, week_text = week_key.split("-W", 1)
    return date.fromisocalendar(int(year_text), int(week_text), 1)


def week_keys_between(start: date, end: date) -> list[str]:
    current = start - timedelta(days=start.weekday())
    final = end - timedelta(days=end.weekday())
    keys: list[str] = []
    while current <= final:
        keys.append(week_key_from_date(current))
        current += timedelta(days=7)
    return keys


def visible_sub_tasks(db: Session, user: User) -> list[SubTask]:
    return [
        task
        for task in db.scalars(select(SubTask).order_by(SubTask.code)).all()
        if task.department_task
        and task.department_task.status != "archived"
        and task.department_task.parent_task
        and task.department_task.parent_task.status != "archived"
        and can_access_sub_task(db, user, task)
    ]


def serialize_sub_task_for_execution(task: SubTask, user: User, current_update: WeeklyUpdate | None = None) -> dict:
    relation = sub_task_execution_relation(user, task)
    executor_ids = sub_task_executor_ids(task)
    assignee_id = user.id if user.id in executor_ids else (task.executor_id if task.executor_id in executor_ids else None)
    owner_options = risk_owner_options(task)
    return {
        **serialize_sub_task(task, current_update),
        "viewer_relation": relation,
        "can_update_weekly": can_update_sub_task_weekly(user, task, assignee_id),
        "can_reopen": can_reopen_sub_task(user, task),
        "current_assignee_id": assignee_id,
        "can_create_risk": can_create_risk_item(user, task),
        "risk_owner_options": owner_options,
        "default_risk_owner_id": owner_options[0]["id"] if owner_options else None,
    }


def visible_department_tasks(db: Session, user: User) -> list[DepartmentTask]:
    return [
        task
        for task in db.scalars(
            select(DepartmentTask)
            .join(ParentTask)
            .where(DepartmentTask.status != "archived", ParentTask.status != "archived")
            .order_by(DepartmentTask.code)
        ).all()
        if can_access_department_task(db, user, task)
    ]


def update_by_sub_task(db: Session, sub_tasks: list[SubTask], week_key: str) -> dict[int, WeeklyUpdate]:
    ids = [task.id for task in sub_tasks]
    if not ids:
        return {}
    return {
        update.sub_task_id: update
        for update in db.scalars(
            select(WeeklyUpdate).where(WeeklyUpdate.week_key == week_key, WeeklyUpdate.sub_task_id.in_(ids))
        ).all()
    }


def earliest_update_week_map(db: Session, sub_tasks: list[SubTask]) -> dict[int, str]:
    ids = [task.id for task in sub_tasks]
    if not ids:
        return {}
    result: dict[int, str] = {}
    for update in db.scalars(
        select(WeeklyUpdate).where(WeeklyUpdate.sub_task_id.in_(ids)).order_by(WeeklyUpdate.week_key)
    ).all():
        result.setdefault(update.sub_task_id, update.week_key)
    return result


def task_start_date(task: SubTask, earliest_weeks: dict[int, str]) -> date:
    if task.started_at:
        return task.started_at.date()
    if task.id in earliest_weeks:
        return monday_from_week_key(earliest_weeks[task.id])
    return monday_from_week_key(current_week_key())


def sub_task_summary(task: SubTask, current_update: WeeklyUpdate | None = None) -> dict:
    return {
        **serialize_sub_task(task, current_update),
        "parent_task_id": task.department_task.parent_task_id if task.department_task else None,
        "parent_task_code": task.department_task.parent_task.code if task.department_task and task.department_task.parent_task else None,
        "parent_task": task.department_task.parent_task.title if task.department_task and task.department_task.parent_task else None,
        "department_task_code": task.department_task.code if task.department_task else None,
        "department_task": task.department_task.title if task.department_task else None,
        "started_at": task.started_at.isoformat() if task.started_at else None,
    }


def sub_task_detail_summary(task: SubTask, updates: list[WeeklyUpdate], current_update: WeeklyUpdate | None = None) -> dict:
    return {
        **sub_task_summary(task, current_update or (updates[0] if updates else None)),
        **weekly_update_summary(updates),
    }


def parent_task_detail_summary(task: ParentTask) -> dict:
    owner_ids, owners, owner_text = people_payload(owner_people(task))
    return {
        "id": task.id,
        "code": task.code,
        "title": task.title,
        "owner": owner_text,
        "owner_ids": owner_ids,
        "owners": owners,
        "department": task.department.name if task.department else None,
        "status": task.status,
        "due_date": task.due_date.isoformat() if task.due_date else None,
    }


def build_meeting_overview_payload(db: Session, user: User, week_key: str) -> dict:
    sub_tasks = visible_sub_tasks(db, user)
    current_updates = update_by_sub_task(db, sub_tasks, week_key)
    grouped_updates = updates_by_sub_task(db, [task.id for task in sub_tasks], week_key)
    today = date.today()
    active_tasks = [task for task in sub_tasks if task.status != "completed"]
    updated_tasks = [task for task in active_tasks if task.id in current_updates]
    missing_tasks = [task for task in active_tasks if task.id not in current_updates]
    risk_items = active_risk_items_for_tasks(db, sub_tasks)
    high_risks = [item for item in risk_items if item.level == "high"]
    medium_risks = [item for item in risk_items if item.level == "medium"]
    low_risks = [item for item in risk_items if item.level == "low"]
    overdue_tasks = [task for task in active_tasks if task.due_date and task.due_date < today]
    completed_tasks = [task for task in sub_tasks if task.status == "completed"]
    weeks = week_keys_between(monday_from_week_key(week_key) - timedelta(days=35), monday_from_week_key(week_key))
    trend_updates = list(
        db.scalars(select(WeeklyUpdate).where(WeeklyUpdate.week_key.in_(weeks))).all()
    )
    visible_ids = {task.id for task in sub_tasks}
    trend = [
        {
            "week_key": key,
            "submitted": len([item for item in trend_updates if item.week_key == key and item.status == "submitted" and item.sub_task_id in visible_ids]),
            "draft": len([item for item in trend_updates if item.week_key == key and item.status == "draft" and item.sub_task_id in visible_ids]),
        }
        for key in weeks
    ]
    visible_parent_tasks = {
        task.department_task.parent_task.id: task.department_task.parent_task
        for task in sub_tasks
        if task.department_task and task.department_task.parent_task
    }
    gantt_source = sorted(visible_parent_tasks.values(), key=lambda item: item.due_date or date.max)[:18]

    def details_for(tasks: list[SubTask]) -> list[dict]:
        return [
            sub_task_detail_summary(task, grouped_updates.get(task.id, []), current_updates.get(task.id))
            for task in tasks
        ]

    detail_rows = {
        "active_sub_tasks": details_for(active_tasks),
        "updated_this_week": details_for(updated_tasks),
        "missing_updates": details_for(missing_tasks),
        "risk_tasks": [risk_item_detail_summary(item, user) for item in risk_items],
        "overdue_tasks": details_for(overdue_tasks),
        "completed_tasks": details_for(completed_tasks),
        "weekly_updated": details_for(updated_tasks),
        "weekly_missing": details_for(missing_tasks),
        "weekly_completed": details_for(completed_tasks),
        "risk_high": [risk_item_detail_summary(item, user) for item in high_risks],
        "risk_medium": [risk_item_detail_summary(item, user) for item in medium_risks],
        "risk_low": [risk_item_detail_summary(item, user) for item in low_risks],
    }
    return {
        "week_key": week_key,
        "cards": {
            "active_sub_tasks": len(active_tasks),
            "updated_this_week": len(updated_tasks),
            "missing_updates": len(missing_tasks),
            "risk_tasks": len(risk_items),
            "overdue_tasks": len(overdue_tasks),
            "completed_tasks": len(completed_tasks),
        },
        "weekly_bar": [
            {"name": "已更新", "value": len(updated_tasks), "detail_key": "weekly_updated"},
            {"name": "待更新", "value": len(missing_tasks), "detail_key": "weekly_missing"},
            {"name": "已完成", "value": len(completed_tasks), "detail_key": "weekly_completed"},
        ],
        "risk_pie": [
            {"name": "高风险", "value": len(detail_rows["risk_high"]), "detail_key": "risk_high"},
            {"name": "中风险", "value": len(detail_rows["risk_medium"]), "detail_key": "risk_medium"},
            {"name": "低风险", "value": len(detail_rows["risk_low"]), "detail_key": "risk_low"},
        ],
        "trend": trend,
        "gantt": [
            {
                "id": task.id,
                "code": task.code,
                "title": task.title,
                "owner": people_payload(owner_people(task))[2],
                "department": task.department.name if task.department else None,
                "status": task.status,
                "start_date": task.due_date.replace(day=1).isoformat() if task.due_date else today.replace(day=1).isoformat(),
                "due_date": task.due_date.isoformat() if task.due_date else today.isoformat(),
            }
            for task in gantt_source
        ],
        "details": detail_rows,
        "parent_details": [parent_task_detail_summary(task) for task in gantt_source],
        "risk_overdue": [
            *[risk_item_detail_summary(item, user) for item in sorted(risk_items, key=lambda item: item.due_date or date.max)],
            *[
            {
                **sub_task_detail_summary(task, grouped_updates.get(task.id, []), current_updates.get(task.id)),
                "issue_type": "逾期",
            }
            for task in sorted(overdue_tasks, key=lambda item: item.due_date or date.max)
            ],
        ],
    }


def build_parent_board_payload(db: Session, user: User, week_key: str) -> dict:
    sub_tasks = visible_sub_tasks(db, user)
    current_updates = update_by_sub_task(db, sub_tasks, week_key)
    risk_items = active_risk_items_for_tasks(db, sub_tasks)
    today = date.today()
    rows = []
    by_parent: dict[int, list[SubTask]] = defaultdict(list)
    for task in sub_tasks:
        if task.department_task and task.department_task.parent_task:
            by_parent[task.department_task.parent_task.id].append(task)
    for parent_id, tasks in by_parent.items():
        parent = tasks[0].department_task.parent_task
        department_task_ids = {task.department_task_id for task in tasks}
        active = [task for task in tasks if task.status != "completed"]
        owner_ids, owners, owner_text = people_payload(owner_people(parent))
        rows.append(
            {
                "id": parent_id,
                "code": parent.code,
                "title": parent.title,
                "owner": owner_text,
                "owner_ids": owner_ids,
                "owners": owners,
                "department": parent.department.name if parent.department else None,
                "department_task_count": len(department_task_ids),
                "sub_task_count": len(tasks),
                "missing_updates": len([task for task in active if task.id not in current_updates]),
                "risk_count": len(
                    [
                        item
                        for item in risk_items
                        if item.sub_task
                        and item.sub_task.department_task
                        and item.sub_task.department_task.parent_task_id == parent_id
                    ]
                ),
                "overdue_count": len([task for task in active if task.due_date and task.due_date < today]),
                "completed_count": len([task for task in tasks if task.status == "completed"]),
            }
        )
    return {"week_key": week_key, "rows": sorted(rows, key=lambda item: item["code"])}


def build_department_board_payload(db: Session, user: User, week_key: str) -> dict:
    sub_tasks = visible_sub_tasks(db, user)
    current_updates = update_by_sub_task(db, sub_tasks, week_key)
    risk_items = active_risk_items_for_tasks(db, sub_tasks)
    risk_counts_by_department: dict[int, int] = defaultdict(int)
    for item in risk_items:
        task = item.sub_task
        if not task or not task.department_task:
            continue
        departments = task.department_task.departments or (
            [task.department_task.department] if task.department_task.department else []
        )
        for department in departments:
            risk_counts_by_department[department.id] += 1
    today = date.today()
    rows_by_department: dict[int, dict] = {}
    for task in sub_tasks:
        department_task = task.department_task
        departments = department_task.departments or ([department_task.department] if department_task.department else [])
        for department in departments:
            row = rows_by_department.setdefault(
                department.id,
                {
                    "id": department.id,
                    "name": department.name,
                    "department_task_ids": set(),
                    "sub_task_ids": set(),
                    "missing_updates": 0,
                    "risk_count": 0,
                    "overdue_count": 0,
                    "completed_count": 0,
                },
            )
            row["department_task_ids"].add(department_task.id)
            row["sub_task_ids"].add(task.id)
            if task.status != "completed" and task.id not in current_updates:
                row["missing_updates"] += 1
            if task.status != "completed" and task.due_date and task.due_date < today:
                row["overdue_count"] += 1
            if task.status == "completed":
                row["completed_count"] += 1
    rows = []
    for row in rows_by_department.values():
        rows.append(
            {
                **{key: value for key, value in row.items() if key not in {"department_task_ids", "sub_task_ids"}},
                "risk_count": risk_counts_by_department.get(row["id"], 0),
                "department_task_count": len(row["department_task_ids"]),
                "sub_task_count": len(row["sub_task_ids"]),
            }
        )
    return {"week_key": week_key, "rows": sorted(rows, key=lambda item: item["name"])}


def build_timeline_matrix_payload(db: Session, user: User) -> dict:
    sub_tasks = visible_sub_tasks(db, user)
    current_monday = monday_from_week_key(current_week_key())
    earliest_weeks = earliest_update_week_map(db, sub_tasks)
    start_dates = [task_start_date(task, earliest_weeks) for task in sub_tasks] or [current_monday]
    first_monday = min(start_dates) - timedelta(days=min(start_dates).weekday())
    if (current_monday - first_monday).days > 49:
        first_monday = current_monday - timedelta(days=49)
    weeks = week_keys_between(first_monday, current_monday)
    updates = {
        (update.sub_task_id, update.week_key): update
        for update in db.scalars(
            select(WeeklyUpdate).where(
                WeeklyUpdate.sub_task_id.in_([task.id for task in sub_tasks] or [0]),
                WeeklyUpdate.week_key.in_(weeks),
            )
        ).all()
    }
    attachments_by_update: dict[int, list[dict]] = defaultdict(list)
    update_ids = [update.id for update in updates.values()]
    if update_ids:
        for attachment in db.scalars(
            select(Attachment).where(
                Attachment.related_type == "weekly_update",
                Attachment.related_id.in_(update_ids),
            )
        ).all():
            if attachment_file_exists(attachment):
                attachments_by_update[attachment.related_id].append(
                    {"id": attachment.id, "filename": attachment.filename}
                )
    tree: dict[int, dict] = {}
    for task in sub_tasks:
        department_task = task.department_task
        parent_task = department_task.parent_task
        parent_node = tree.setdefault(
            parent_task.id,
            {"id": parent_task.id, "code": parent_task.code, "title": parent_task.title, "department_tasks": {}},
        )
        department_node = parent_node["department_tasks"].setdefault(
            department_task.id,
            {"id": department_task.id, "code": department_task.code, "title": department_task.title, "sub_tasks": []},
        )
        cells = {}
        for week in weeks:
            update = updates.get((task.id, week))
            cells[week] = {
                "this_week": update.this_week if update else None,
                "risk": update.risk if update else None,
                "attachments": attachments_by_update.get(update.id, []) if update else [],
            }
        department_node["sub_tasks"].append(
            {
                "id": task.id,
                "code": task.code,
                "title": task.title,
                "executor": people_payload(executor_people(task))[2],
                "owner": people_payload(owner_people(task))[2],
                "status": task.status,
                "started_at": task_start_date(task, earliest_weeks).isoformat(),
                "cells": cells,
            }
        )
    parents = []
    for parent in sorted(tree.values(), key=lambda item: item["code"]):
        parent["department_tasks"] = sorted(parent["department_tasks"].values(), key=lambda item: item["code"])
        for department_task in parent["department_tasks"]:
            department_task["sub_tasks"] = sorted(department_task["sub_tasks"], key=lambda item: item["code"])
        parents.append(parent)
    return {"week_key": current_week_key(), "weeks": weeks, "parents": parents}


@router.get("/health")
def health() -> dict:
    return {"status": "ok", "version": __version__}


@router.get("/auth/me")
def me(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)) -> dict:
    guides = guides_payload(db, current_user)
    return {
        "user": serialize_user(current_user),
        "permission_codes": sorted(user_permission_codes(current_user)),
        "week_key": current_week_key(),
        "features": feature_payload(db, current_user),
        "onboarding": onboarding_payload(current_user),
        "guide_profile": guides["profile"],
        "guides": {
            "system": guides["system"],
            "modules": guides["modules"],
        },
    }


@router.post("/auth/onboarding")
def update_onboarding(
    payload: OnboardingUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> dict:
    if payload.version != CURRENT_ONBOARDING_VERSION:
        raise HTTPException(status.HTTP_409_CONFLICT, detail="使用指南版本已更新，请刷新后重试")
    if payload.action not in {"completed", "skipped"}:
        raise HTTPException(status.HTTP_422_UNPROCESSABLE_ENTITY, detail="无效的使用指南状态")
    current_user.onboarding_version = CURRENT_ONBOARDING_VERSION
    if current_user.onboarding_status != "completed" or payload.action == "completed":
        current_user.onboarding_status = payload.action
    current_user.onboarding_completed_at = datetime.now(timezone.utc)
    db.add(current_user)
    db.commit()
    return onboarding_payload(current_user)


@router.post("/auth/guides")
def update_guide_progress(
    payload: GuideProgressUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> dict:
    if payload.action not in {"completed", "skipped"}:
        raise HTTPException(status.HTTP_422_UNPROCESSABLE_ENTITY, detail="无效的使用指南状态")
    if (payload.guide_key, payload.version) not in allowed_guides(db, current_user):
        raise HTTPException(status.HTTP_403_FORBIDDEN, detail="当前用户不可更新该使用指南")
    progress = db.scalar(
        select(UserGuideProgress).where(
            UserGuideProgress.user_id == current_user.id,
            UserGuideProgress.guide_key == payload.guide_key,
            UserGuideProgress.version == payload.version,
        )
    )
    if progress:
        if progress.status != "completed" or payload.action == "completed":
            progress.status = payload.action
        progress.completed_at = datetime.now(timezone.utc)
    else:
        progress = UserGuideProgress(
            user_id=current_user.id,
            guide_key=payload.guide_key,
            version=payload.version,
            status=payload.action,
            completed_at=datetime.now(timezone.utc),
        )
        db.add(progress)
    db.commit()
    return guides_payload(db, current_user)


@router.post("/auth/login")
def login(payload: LoginRequest, response: Response, db: Session = Depends(get_db)) -> dict:
    user = db.scalar(select(User).where(User.username == payload.username))
    if not user or not user.password_hash or not verify_password(payload.password, user.password_hash):
        raise HTTPException(status.HTTP_401_UNAUTHORIZED, detail="用户名或密码错误")
    if user.status not in {"active", "pending"}:
        raise HTTPException(status.HTTP_403_FORBIDDEN, detail="账号已停用")
    token = create_session(db, user)
    response.set_cookie(
        SESSION_COOKIE,
        token,
        **session_cookie_kwargs(),
    )
    db.refresh(user)
    return {
        "user": serialize_user(user),
        "permission_codes": sorted(user_permission_codes(user)),
        "week_key": current_week_key(),
        "features": feature_payload(db, user),
    }


@router.post("/auth/logout")
def logout(
    response: Response,
    db: Session = Depends(get_db),
    session_token: str | None = Cookie(default=None, alias=SESSION_COOKIE),
) -> dict:
    delete_session(db, session_token)
    response.delete_cookie(SESSION_COOKIE)
    return {"ok": True}


@router.post("/auth/openid-login")
def openid_login(payload: OpenIdLoginRequest, response: Response, db: Session = Depends(get_db)) -> dict:
    user = db.scalar(select(User).where(User.open_id == payload.open_id))
    if not user:
        user = db.scalar(select(User).where(User.open_id.is_(None), User.name == payload.name))
        if user:
            user.open_id = payload.open_id
            user.open_id_bound_at = datetime.now(timezone.utc)
            user.source = user.source or "manual"
        else:
            user = User(
                name=payload.name,
                open_id=payload.open_id,
                status="pending",
                source="feishu_pending",
                open_id_bound_at=datetime.now(timezone.utc),
            )
            db.add(user)
    if user.status == "disabled":
        raise HTTPException(status.HTTP_403_FORBIDDEN, detail="账号已停用")
    db.add(user)
    db.commit()
    db.refresh(user)
    token = create_session(db, user)
    response.set_cookie(
        SESSION_COOKIE,
        token,
        **session_cookie_kwargs(),
    )
    return {
        "user": serialize_user(user),
        "permission_codes": sorted(user_permission_codes(user)),
        "week_key": current_week_key(),
        "features": feature_payload(db, user),
    }


@router.get("/auth/lark-link")
def lark_link_login(token: str, db: Session = Depends(get_db)) -> RedirectResponse:
    user, next_path, notification = verify_lark_login_token(db, token)
    if notification:
        clicked_at = datetime.now(timezone.utc)
        db.execute(
            update(NotificationRecord)
            .where(NotificationRecord.id == notification.id)
            .values(
                clicked=True,
                first_clicked_at=func.coalesce(NotificationRecord.first_clicked_at, clicked_at),
                last_clicked_at=clicked_at,
                click_count=func.coalesce(NotificationRecord.click_count, 0) + 1,
            )
        )
    session_token = create_session(db, user)
    redirect = RedirectResponse(url=next_path, status_code=status.HTTP_302_FOUND)
    redirect.set_cookie(
        SESSION_COOKIE,
        session_token,
        **session_cookie_kwargs(),
    )
    return redirect


@router.get("/auth/lark-oauth/start")
def lark_oauth_start(request: Request, next_path: str | None = None) -> RedirectResponse:
    try:
        url = create_lark_oauth_authorize_url(
            next_path,
            request_host=request.headers.get("host"),
            request_scheme=request.headers.get("x-forwarded-proto") or request.url.scheme,
        )
    except RuntimeError as exc:
        raise HTTPException(status.HTTP_503_SERVICE_UNAVAILABLE, detail=str(exc)) from exc
    return RedirectResponse(url=url, status_code=status.HTTP_302_FOUND)


@router.get("/auth/lark-oauth/callback")
async def lark_oauth_callback(
    code: str | None = None,
    state: str | None = None,
    db: Session = Depends(get_db),
) -> RedirectResponse:
    if not code or not state:
        return lark_login_error_redirect("飞书免登参数缺失，请联系管理员")
    try:
        next_path = verify_lark_oauth_state(state)
        token_payload = await lark_client.get_user_access_token(code)
        user_access_token = token_payload.get("access_token") or token_payload.get("user_access_token")
        user_info = await lark_client.get_user_info(user_access_token) if user_access_token else {}
        open_id = normalize_open_id(user_info.get("open_id") or token_payload.get("open_id"))
        email = normalize_email(user_info.get("email") or token_payload.get("email"))
        if not open_id:
            return lark_login_error_redirect("飞书免登未返回 open_id，请联系管理员检查权限")

        user = db.scalar(select(User).where(User.open_id == open_id))
        if not user and email:
            matches = list(db.scalars(select(User).where(User.email == email)).all())
            if len(matches) == 1:
                user = matches[0]
                conflict = db.scalar(select(User).where(User.open_id == open_id, User.id != user.id))
                if conflict:
                    return lark_login_error_redirect("飞书身份已绑定到其他人员，请联系管理员")
                user.open_id = open_id
                user.open_id_bound_at = datetime.now(timezone.utc)
                db.add(user)
                db.commit()
                db.refresh(user)
            elif len(matches) > 1:
                return lark_login_error_redirect("邮箱匹配到多个人员，请联系管理员处理")
        if not user:
            return lark_login_error_redirect("未找到对应人员档案，请联系管理员绑定")
        if user.status == "disabled":
            return lark_login_error_redirect("账号已停用，请联系管理员")
        session_token = create_session(db, user)
    except HTTPException as exc:
        return lark_login_error_redirect(str(exc.detail))
    except Exception as exc:
        return lark_login_error_redirect(f"飞书免登失败：{exc}")

    redirect = RedirectResponse(url=next_path, status_code=status.HTTP_302_FOUND)
    redirect.set_cookie(
        SESSION_COOKIE,
        session_token,
        **session_cookie_kwargs(),
    )
    return redirect


def serialize_person(user: User) -> dict:
    item = serialize_user(user) or {}
    item.update(
        {
            "username": user.username,
            "status": user.status,
            "source": user.source,
            "is_admin": user.is_admin,
            "email": user.email,
            "open_id_bound_at": user.open_id_bound_at.isoformat() if user.open_id_bound_at else None,
            "last_login_at": user.last_login_at.isoformat() if user.last_login_at else None,
        }
    )
    return item


@router.get("/users")
def list_users(db: Session = Depends(get_db), _: User = Depends(require_admin)) -> list[dict]:
    return [serialize_user(user) for user in db.scalars(select(User).order_by(User.id)).all()]


@router.get("/user-options")
def list_user_options(db: Session = Depends(get_db), _: User = Depends(get_current_user)) -> list[dict]:
    return [
        {
            "id": user.id,
            "name": user.name,
            "department": user.department.name if user.department else None,
            "title": user.title,
        }
        for user in db.scalars(select(User).where(User.status != "disabled").order_by(User.name)).all()
    ]


@router.get("/people")
def list_people(
    db: Session = Depends(get_db),
    _: User = Depends(require_admin),
) -> list[dict]:
    users = db.scalars(select(User).order_by(User.is_admin.desc(), User.id)).all()
    return [serialize_person(user) for user in users]


@router.post("/people", status_code=201)
def create_person(
    payload: PersonCreate,
    db: Session = Depends(get_db),
    _: User = Depends(require_admin),
) -> dict:
    open_id = normalize_open_id(payload.open_id)
    email = normalize_email(payload.email)
    if open_id and db.scalar(select(User).where(User.open_id == open_id)):
        raise HTTPException(status.HTTP_409_CONFLICT, detail="open_id 已绑定到其他人员")
    if email and db.scalar(select(User).where(User.email == email)):
        raise HTTPException(status.HTTP_409_CONFLICT, detail="邮箱已绑定到其他人员")
    if db.scalar(select(User).where(User.name == payload.name, User.open_id.is_(None))):
        raise HTTPException(status.HTTP_409_CONFLICT, detail="已存在同名未绑定人员")
    user = User(
        name=payload.name,
        department_id=payload.department_id,
        title=payload.title,
        status=payload.status,
        source="manual",
        open_id=open_id,
        email=email,
        open_id_bound_at=datetime.now(timezone.utc) if open_id else None,
    )
    roles = db.scalars(select(Role).where(Role.id.in_(payload.role_ids))).all() if payload.role_ids else []
    user.roles = list(roles)
    db.add(user)
    db.commit()
    db.refresh(user)
    return serialize_person(user)


@router.put("/people/{user_id}")
def update_person(
    user_id: int,
    payload: PersonUpdate,
    db: Session = Depends(get_db),
    _: User = Depends(require_admin),
) -> dict:
    user = db.get(User, user_id)
    if not user:
        raise HTTPException(status.HTTP_404_NOT_FOUND, detail="User not found")
    data = payload.model_dump(exclude_unset=True)
    role_ids = data.pop("role_ids", None)
    open_id = data.pop("open_id", None)
    email = data.pop("email", None)
    for key, value in data.items():
        setattr(user, key, value)
    if "open_id" in payload.model_fields_set:
        apply_open_id_binding(db, user, open_id)
    if "email" in payload.model_fields_set:
        apply_email_binding(db, user, email)
    if role_ids is not None:
        user.roles = list(db.scalars(select(Role).where(Role.id.in_(role_ids))).all()) if role_ids else []
    db.add(user)
    db.commit()
    db.refresh(user)
    return serialize_person(user)


def parse_email_import(content: bytes, filename: str) -> list[dict[str, str]]:
    lower_name = filename.lower()
    rows: list[dict[str, str]] = []
    if lower_name.endswith(".csv"):
        text = content.decode("utf-8-sig")
        reader = csv.DictReader(StringIO(text))
        for row in reader:
            rows.append({str(key or "").strip(): str(value or "").strip() for key, value in row.items()})
        return rows
    if lower_name.endswith(".xlsx"):
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
        worksheet = workbook.active
        header = [str(value or "").strip() for value in next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))]
        for values in worksheet.iter_rows(min_row=2, values_only=True):
            rows.append(
                {
                    header[index]: str(value or "").strip()
                    for index, value in enumerate(values)
                    if index < len(header)
                }
            )
        return rows
    raise HTTPException(status.HTTP_422_UNPROCESSABLE_ENTITY, detail="仅支持 CSV 或 XLSX 邮箱导入文件")


def normalize_department_name(name: str) -> str:
    normalized = re.sub(r"\s+", " ", (name or "").strip())
    if not normalized:
        raise HTTPException(status.HTTP_422_UNPROCESSABLE_ENTITY, detail="部门名称不能为空")
    if len(normalized) > 120:
        raise HTTPException(status.HTTP_422_UNPROCESSABLE_ENTITY, detail="部门名称不能超过 120 个字符")
    return normalized


def department_reference_counts(db: Session, department_id: int) -> dict[str, int]:
    return {
        "users": db.scalar(select(func.count()).select_from(User).where(User.department_id == department_id)) or 0,
        "parent_tasks": db.scalar(select(func.count()).select_from(ParentTask).where(ParentTask.department_id == department_id)) or 0,
        "department_tasks": db.scalar(select(func.count()).select_from(DepartmentTask).where(DepartmentTask.department_id == department_id)) or 0,
        "department_task_departments": db.scalar(
            select(func.count()).select_from(DepartmentTaskDepartment).where(DepartmentTaskDepartment.department_id == department_id)
        )
        or 0,
        "child_departments": db.scalar(select(func.count()).select_from(Department).where(Department.parent_id == department_id)) or 0,
    }


def serialize_department_item(item: Department, reference_counts: dict[str, int] | None = None) -> dict:
    payload = {
        "id": item.id,
        "name": item.name,
        "manager_id": item.manager_id,
        "manager": item.manager.name if item.manager else None,
        "status": item.status,
    }
    if reference_counts is not None:
        blocking = {
            "人员": reference_counts["users"],
            "母任务": reference_counts["parent_tasks"],
            "部门任务": reference_counts["department_tasks"],
            "部门任务多部门关联": reference_counts["department_task_departments"],
            "子部门": reference_counts["child_departments"],
        }
        blocking_reasons = [f"{label} {count} 项" for label, count in blocking.items() if count]
        payload.update(
            {
                "reference_counts": reference_counts,
                "reference_total": sum(reference_counts.values()),
                "can_delete": not blocking_reasons,
                "delete_blocking_reasons": blocking_reasons,
            }
        )
    return payload


@router.post("/lark/import-user-emails")
async def import_user_emails(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    _: User = Depends(require_admin),
) -> dict:
    content = await file.read()
    rows = parse_email_import(content, file.filename or "")
    summary = sync_people_from_rows(db, rows, apply=True)
    summary["open_id_candidates_after_apply"] = count_open_id_candidates(db)
    db.commit()
    return {
        "ok": summary["blocked"] == 0,
        "scanned": summary["source_rows"],
        "imported": summary["email_updates"],
        "created": summary["people_created"],
        "unchanged": summary["people_unchanged"],
        "skipped": summary["skipped_company_contacts"],
        "blocked": summary["blocked"],
        "open_id_candidates": summary["open_id_candidates_after_apply"],
        "results": summary["results"],
    }


@router.get("/departments")
def list_departments(db: Session = Depends(get_db), _: User = Depends(get_current_user)) -> list[dict]:
    departments = db.scalars(select(Department).order_by(Department.id)).all()
    return [serialize_department_item(item) for item in departments]


@router.get("/departments/manage")
def manage_departments(db: Session = Depends(get_db), _: User = Depends(require_admin)) -> list[dict]:
    departments = db.scalars(select(Department).order_by(Department.id)).all()
    return [serialize_department_item(item, department_reference_counts(db, item.id)) for item in departments]


@router.post("/departments", status_code=201)
def create_department(
    payload: DepartmentCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin),
) -> dict:
    name = normalize_department_name(payload.name)
    if db.scalar(select(Department).where(Department.name == name)):
        raise HTTPException(status.HTTP_409_CONFLICT, detail="部门名称已存在")
    department = Department(name=name, status="active")
    db.add(department)
    db.flush()
    db.add(
        TaskEvent(
            object_type="department",
            object_id=department.id,
            event_type="department_created",
            title="新增部门",
            content=name,
            actor_id=current_user.id,
        )
    )
    db.commit()
    db.refresh(department)
    return serialize_department_item(department, department_reference_counts(db, department.id))


@router.put("/departments/{department_id}")
def update_department(
    department_id: int,
    payload: DepartmentUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin),
) -> dict:
    department = db.get(Department, department_id)
    if not department:
        raise HTTPException(status.HTTP_404_NOT_FOUND, detail="Department not found")
    name = normalize_department_name(payload.name)
    existing = db.scalar(select(Department).where(Department.name == name, Department.id != department_id))
    if existing:
        raise HTTPException(status.HTTP_409_CONFLICT, detail="部门名称已存在")
    old_name = department.name
    department.name = name
    db.add(
        TaskEvent(
            object_type="department",
            object_id=department.id,
            event_type="department_renamed",
            title="编辑部门",
            content=f"{old_name} -> {name}",
            actor_id=current_user.id,
        )
    )
    db.commit()
    db.refresh(department)
    return serialize_department_item(department, department_reference_counts(db, department.id))


@router.delete("/departments/{department_id}")
def delete_department(
    department_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin),
) -> dict:
    department = db.get(Department, department_id)
    if not department:
        raise HTTPException(status.HTTP_404_NOT_FOUND, detail="Department not found")
    reference_counts = department_reference_counts(db, department.id)
    if sum(reference_counts.values()):
        detail = serialize_department_item(department, reference_counts)
        detail["message"] = "该部门仍有引用，不能删除"
        raise HTTPException(status.HTTP_409_CONFLICT, detail=detail)
    db.add(
        TaskEvent(
            object_type="department",
            object_id=department.id,
            event_type="department_deleted",
            title="删除部门",
            content=department.name,
            actor_id=current_user.id,
        )
    )
    db.delete(department)
    db.commit()
    return {"ok": True, "id": department_id}


@router.get("/roles")
def list_roles(db: Session = Depends(get_db), _: User = Depends(get_current_user)) -> list[dict]:
    return [
        {
            "id": role.id,
            "code": role.code,
            "name": role.name,
            "description": role.description,
            "permission_codes": [permission.code for permission in role.permissions],
        }
        for role in db.scalars(select(Role).order_by(Role.id)).all()
    ]


@router.get("/permissions")
def list_permissions(db: Session = Depends(get_db), _: User = Depends(require_admin)) -> dict:
    roles = db.scalars(select(Role).order_by(Role.id)).all()
    permissions = db.scalars(select(Permission).order_by(Permission.id)).all()
    return {
        "permissions": [
            {"id": item.id, "code": item.code, "name": item.name, "description": item.description}
            for item in permissions
        ],
        "matrix": [
            {
                "role_id": role.id,
                "role_code": role.code,
                "role_name": role.name,
                "permission_codes": [permission.code for permission in role.permissions],
            }
            for role in roles
        ],
    }


@router.put("/permissions/matrix")
def update_permission_matrix(
    payload: RolePermissionUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin),
) -> dict:
    role = db.get(Role, payload.role_id)
    if not role:
        raise HTTPException(status.HTTP_404_NOT_FOUND, detail="Role not found")
    refresh_role_permissions(db, role, payload.permission_codes)
    db.add(
        TaskEvent(
            object_type="role",
            object_id=role.id,
            event_type="permission_matrix_updated",
            title="更新权限矩阵",
            content=",".join(payload.permission_codes),
            actor_id=current_user.id,
        )
    )
    db.commit()
    return {"ok": True}


@router.post("/sync/base-2026/preview")
def preview_base_sync(_: User = Depends(require_admin)) -> dict:
    return preview_base_2026()


@router.post("/sync/base-2026/import")
def import_base_sync(
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin),
) -> dict:
    return import_base_2026(db, current_user.id)


@router.get("/goals")
def list_goals(db: Session = Depends(get_db), _: User = Depends(get_current_user)) -> list[dict]:
    return [serialize_goal(goal) for goal in db.scalars(select(StrategicGoal).order_by(StrategicGoal.id))]


@router.get("/goals/{goal_id}/parent-tasks")
def list_goal_parent_tasks(
    goal_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> list[dict]:
    goal = db.get(StrategicGoal, goal_id)
    if not goal:
        raise HTTPException(status.HTTP_404_NOT_FOUND, detail="Goal not found")
    tasks = db.scalars(
        select(ParentTask)
        .where(ParentTask.goal_id == goal_id, ParentTask.status != "archived")
        .order_by(ParentTask.code)
    ).all()
    return [
        serialize_parent_task_for_user(db, current_user, task)
        for task in tasks
        if can_access_parent_task(current_user, task)
    ]


@router.post("/goals", status_code=201)
def create_goal(
    payload: GoalCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission("task.create_parent")),
) -> dict:
    goal = StrategicGoal(
        code=generate_code(db, StrategicGoal, "SG"),
        name=payload.name,
        description=payload.description,
        year=payload.year,
        progress=payload.progress,
    )
    db.add(goal)
    db.flush()
    db.add(
        TaskEvent(
            object_type="strategic_goal",
            object_id=goal.id,
            event_type="created",
            title="创建战略目标",
            content=goal.name,
            actor_id=current_user.id,
        )
    )
    db.commit()
    db.refresh(goal)
    return serialize_goal(goal)


@router.get("/parent-tasks")
def list_parent_tasks(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)) -> list[dict]:
    if not can_view_parent_task_page(db, current_user):
        raise HTTPException(status.HTTP_403_FORBIDDEN, detail="No access to parent task management")
    return [
        serialize_parent_task_for_user(db, current_user, task)
        for task in db.scalars(
            select(ParentTask).where(ParentTask.status != "archived").order_by(ParentTask.id)
        ).all()
        if can_access_parent_task(current_user, task)
    ]


@router.get("/parent-tasks/{parent_task_id}/department-tasks")
def list_parent_department_tasks(
    parent_task_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> list[dict]:
    parent_task = db.get(ParentTask, parent_task_id)
    if not parent_task or parent_task.status == "archived":
        raise HTTPException(status.HTTP_404_NOT_FOUND, detail="Parent task not found")
    if not can_access_parent_task(current_user, parent_task):
        raise HTTPException(status.HTTP_403_FORBIDDEN, detail="No access to this parent task")
    return [
        serialize_department_task_tree_for_user(db, current_user, task)
        for task in db.scalars(
            select(DepartmentTask)
            .join(ParentTask)
            .where(DepartmentTask.parent_task_id == parent_task_id)
            .where(ParentTask.status != "archived")
            .where(DepartmentTask.status != "archived")
            .order_by(DepartmentTask.code)
        ).all()
        if can_access_department_task(db, current_user, task)
    ]


@router.get("/parent-tasks/{parent_task_id}")
def get_parent_task(
    parent_task_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> dict:
    parent_task = db.get(ParentTask, parent_task_id)
    if not parent_task or parent_task.status == "archived":
        raise HTTPException(status.HTTP_404_NOT_FOUND, detail="Parent task not found")
    if not can_access_parent_task(current_user, parent_task):
        raise HTTPException(status.HTTP_403_FORBIDDEN, detail="No access to this parent task")
    return serialize_parent_task_for_user(db, current_user, parent_task)


@router.post("/parent-tasks", status_code=201)
def create_parent_task(
    payload: ParentTaskCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> dict:
    if not can_manage_parent_tasks(current_user):
        raise HTTPException(status.HTTP_403_FORBIDDEN, detail="No create access to parent tasks")
    owners = resolve_users(db, payload.owner_ids, payload.owner_id, "母任务负责人")
    data = payload.model_dump(exclude={"owner_ids"})
    data["owner_id"] = owners[0].id
    task = ParentTask(code=generate_code(db, ParentTask, "MT"), **data)
    db.add(task)
    db.flush()
    sync_task_owners(db, task, owners)
    db.add(
        TaskEvent(
            object_type="parent_task",
            object_id=task.id,
            event_type="created",
            title="创建母任务",
            content=task.title,
            actor_id=current_user.id,
        )
    )
    db.commit()
    db.refresh(task)
    expire_task_people(db, task)
    return serialize_parent_task_for_user(db, current_user, task)


@router.put("/parent-tasks/{parent_task_id}")
def update_parent_task(
    parent_task_id: int,
    payload: ParentTaskUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> dict:
    task = db.get(ParentTask, parent_task_id)
    if not task or task.status == "archived":
        raise HTTPException(status.HTTP_404_NOT_FOUND, detail="Parent task not found")
    if not can_edit_parent_task(current_user, task):
        raise HTTPException(status.HTTP_403_FORBIDDEN, detail="No edit access to this parent task")
    data = payload.model_dump(exclude_unset=True)
    owner_ids = data.pop("owner_ids", None)
    owner_id = data.pop("owner_id", None)
    if owner_ids is not None or owner_id is not None:
        sync_task_owners(db, task, resolve_users(db, owner_ids, owner_id, "母任务负责人"))
    for key, value in data.items():
        setattr(task, key, value)
    db.add(
        TaskEvent(
            object_type="parent_task",
            object_id=task.id,
            event_type="updated",
            title="编辑母任务",
            content=task.title,
            actor_id=current_user.id,
        )
    )
    db.commit()
    db.refresh(task)
    expire_task_people(db, task)
    return serialize_parent_task_for_user(db, current_user, task)


@router.delete("/parent-tasks/{parent_task_id}")
def archive_parent_task(
    parent_task_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> dict:
    task = db.get(ParentTask, parent_task_id)
    if not task or task.status == "archived":
        raise HTTPException(status.HTTP_404_NOT_FOUND, detail="Parent task not found")
    if not can_manage_parent_tasks(current_user):
        raise HTTPException(status.HTTP_403_FORBIDDEN, detail="No delete access to parent tasks")
    task.status = "archived"
    db.add(
        TaskEvent(
            object_type="parent_task",
            object_id=task.id,
            event_type="archived",
            title="归档母任务",
            content=task.title,
            actor_id=current_user.id,
        )
    )
    db.commit()
    return {"ok": True}


@router.get("/department-tasks")
def list_department_tasks(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> list[dict]:
    return [
        serialize_department_task_tree_for_user(db, current_user, task)
        for task in db.scalars(
            select(DepartmentTask)
            .join(ParentTask)
            .where(ParentTask.status != "archived", DepartmentTask.status != "archived")
            .order_by(DepartmentTask.id)
        ).all()
        if can_access_department_task(db, current_user, task)
    ]


@router.get("/department-tasks/overview")
def department_tasks_overview(
    department_id: int | None = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> dict:
    can_switch = can_view_department_directory(current_user)
    all_tasks = [
        task
        for task in db.scalars(
            select(DepartmentTask)
            .join(ParentTask)
            .where(ParentTask.status != "archived", DepartmentTask.status != "archived")
            .order_by(DepartmentTask.code)
        ).all()
        if can_access_department_task(db, current_user, task)
    ]
    visible_department_ids: set[int] = set()
    for task in all_tasks:
        departments = task.departments or ([task.department] if task.department else [])
        visible_department_ids.update(department.id for department in departments)
    if not can_switch and current_user.department_id:
        visible_department_ids = {current_user.department_id}
    selected_department_id = department_id if can_switch and department_id else None
    if not selected_department_id and not can_switch and current_user.department_id:
        selected_department_id = current_user.department_id
    if selected_department_id:
        all_tasks = [
            task
            for task in all_tasks
            if selected_department_id
            in ({department.id for department in task.departments} | {task.department_id})
        ]
    departments = [
        {"id": department.id, "name": department.name}
        for department in db.scalars(select(Department).order_by(Department.name)).all()
        if department.id in visible_department_ids
    ]
    parent_groups = []
    for parent in db.scalars(select(ParentTask).where(ParentTask.status != "archived").order_by(ParentTask.code)).all():
        tasks = [task for task in all_tasks if task.parent_task_id == parent.id]
        if not tasks:
            continue
        sub_tasks = [sub_task for task in tasks for sub_task in task.sub_tasks]
        risk_items = active_risk_items_for_tasks(db, sub_tasks)
        parent_groups.append(
            {
                **serialize_parent_task(parent),
                "department_task_count": len(tasks),
                "sub_task_count": len(sub_tasks),
                "pending_split_count": sum(task.pending_split_count or 0 for task in tasks),
                "risk_count": len(risk_items),
                "department_tasks": [
                    {
                        **serialize_department_task_for_user(db, current_user, task),
                        "sub_tasks": [serialize_sub_task(sub_task) for sub_task in task.sub_tasks],
                    }
                    for task in tasks
                ],
            }
        )
    return {
        "can_switch_department": can_switch,
        "selected_department_id": selected_department_id,
        "departments": departments,
        "department_tasks": [serialize_department_task_tree_for_user(db, current_user, task) for task in all_tasks],
        "parent_tasks": parent_groups,
    }


@router.post("/department-tasks", status_code=201)
async def create_department_task(
    payload: DepartmentTaskCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> dict:
    parent_task = db.get(ParentTask, payload.parent_task_id)
    if not parent_task or parent_task.status == "archived":
        raise HTTPException(status.HTTP_404_NOT_FOUND, detail="Parent task not found")
    if not can_create_department_task(current_user, parent_task):
        raise HTTPException(status.HTTP_403_FORBIDDEN, detail="No create access to department tasks")
    primary_department_id, departments = resolve_departments(db, payload.department_id, payload.department_ids)
    owners = resolve_users(db, payload.owner_ids, payload.owner_id, "负责人")
    data = payload.model_dump(exclude={"department_ids", "owner_ids"})
    data["department_id"] = primary_department_id
    data["owner_id"] = owners[0].id
    task = DepartmentTask(code=generate_code(db, DepartmentTask, "DT"), **data)
    db.add(task)
    db.flush()
    sync_department_task_departments(db, task, departments)
    sync_task_owners(db, task, owners)
    db.add(
        TaskEvent(
            object_type="department_task",
            object_id=task.id,
            event_type="split",
            title="拆分部门任务",
            content=task.title,
            actor_id=current_user.id,
        )
    )
    db.commit()
    db.refresh(task)
    expire_task_people(db, task)
    try:
        await send_department_task_split_notifications(db, task, owner_people(task), event="created")
    except Exception:
        logger.exception("Department task created but split notification failed", extra={"department_task_id": task.id})
    return serialize_department_task_for_user(db, current_user, task)


@router.put("/department-tasks/{department_task_id}")
async def update_department_task(
    department_task_id: int,
    payload: DepartmentTaskUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> dict:
    task = db.get(DepartmentTask, department_task_id)
    if not task or task.status == "archived":
        raise HTTPException(status.HTTP_404_NOT_FOUND, detail="Department task not found")
    if not can_edit_department_task(db, current_user, task):
        raise HTTPException(status.HTTP_403_FORBIDDEN, detail="No edit access to this department task")
    previous_owner_ids = {user.id for user in owner_people(task)}
    data = payload.model_dump(exclude_unset=True)
    department_ids = data.pop("department_ids", None)
    owner_ids = data.pop("owner_ids", None)
    owner_id = data.pop("owner_id", None)
    if department_ids is not None or "department_id" in data:
        primary_department_id, departments = resolve_departments(db, data.get("department_id"), department_ids)
        data["department_id"] = primary_department_id
        sync_department_task_departments(db, task, departments)
    updated_owners: list[User] | None = None
    if owner_ids is not None or owner_id is not None:
        updated_owners = resolve_users(db, owner_ids, owner_id, "负责人")
        sync_task_owners(db, task, updated_owners)
        sync_department_sub_task_owners(db, task, updated_owners)
    for key, value in data.items():
        setattr(task, key, value)
    db.add(
        TaskEvent(
            object_type="department_task",
            object_id=task.id,
            event_type="updated",
            title="编辑部门任务",
            content=task.title,
            actor_id=current_user.id,
        )
    )
    db.commit()
    db.refresh(task)
    expire_task_people(db, task)
    added_owners = [
        user for user in (updated_owners or owner_people(task)) if user.id not in previous_owner_ids
    ]
    if added_owners:
        try:
            await send_department_task_split_notifications(db, task, added_owners, event="owner_added")
        except Exception:
            logger.exception("Department task updated but owner notification failed", extra={"department_task_id": task.id})
    return serialize_department_task_for_user(db, current_user, task)


@router.delete("/department-tasks/{department_task_id}")
def archive_department_task(
    department_task_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> dict:
    task = db.get(DepartmentTask, department_task_id)
    if not task or task.status == "archived":
        raise HTTPException(status.HTTP_404_NOT_FOUND, detail="Department task not found")
    if not can_edit_department_task(db, current_user, task):
        raise HTTPException(status.HTTP_403_FORBIDDEN, detail="No delete access to this department task")
    task.status = "archived"
    db.add(
        TaskEvent(
            object_type="department_task",
            object_id=task.id,
            event_type="archived",
            title="归档部门任务",
            content=task.title,
            actor_id=current_user.id,
        )
    )
    db.commit()
    return {"ok": True}


@router.get("/sub-tasks")
def list_sub_tasks(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> list[dict]:
    week = current_week_key()
    updates = {
        (update.sub_task_id, update.assignee_id): update
        for update in db.scalars(select(WeeklyUpdate).where(WeeklyUpdate.week_key == week)).all()
    }
    rows = []
    for task in db.scalars(select(SubTask).order_by(SubTask.id)).all():
        if (
            can_view_sub_task_execution_entry(current_user, task)
            and task.department_task
            and task.department_task.status != "archived"
            and task.department_task.parent_task
            and task.department_task.parent_task.status != "archived"
        ):
            assignee_id = current_user.id if current_user.id in sub_task_executor_ids(task) else task.executor_id
            rows.append(serialize_sub_task_for_execution(task, current_user, updates.get((task.id, assignee_id))))
    return rows


@router.get("/sub-tasks/{sub_task_id}")
def get_sub_task(
    sub_task_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> dict:
    task = db.get(SubTask, sub_task_id)
    if not task or not task.department_task or task.department_task.status == "archived":
        raise HTTPException(status.HTTP_404_NOT_FOUND, detail="Sub task not found")
    if not can_access_sub_task(db, current_user, task):
        raise HTTPException(status.HTTP_403_FORBIDDEN, detail="No access to this sub task")
    assignee_id = current_user.id if current_user.id in sub_task_executor_ids(task) else task.executor_id
    update = db.scalar(
        select(WeeklyUpdate).where(
            WeeklyUpdate.sub_task_id == task.id,
            WeeklyUpdate.week_key == current_week_key(),
            WeeklyUpdate.assignee_id == assignee_id,
        )
    )
    return serialize_sub_task_for_execution(task, current_user, update)


@router.post("/sub-tasks", status_code=201)
def create_sub_task(
    payload: SubTaskCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> dict:
    department_task = db.get(DepartmentTask, payload.department_task_id)
    if not department_task or department_task.status == "archived":
        raise HTTPException(status.HTTP_404_NOT_FOUND, detail="Department task not found")
    if not can_split_sub_task(db, current_user, department_task):
        raise HTTPException(status.HTTP_403_FORBIDDEN, detail="No split access to this department task")
    owners = owner_people(department_task)
    if not owners:
        raise HTTPException(status.HTTP_422_UNPROCESSABLE_ENTITY, detail="部门任务负责人不能为空")
    validate_inherited_sub_task_owners(payload, owners)
    executors = resolve_users(db, payload.executor_ids, payload.executor_id, "执行人")
    data = payload.model_dump(exclude={"owner_id", "owner_ids", "executor_id", "executor_ids"})
    data["owner_id"] = owners[0].id
    data["executor_id"] = executors[0].id
    task = SubTask(code=generate_sub_task_code(db, department_task), **data)
    db.add(task)
    if department_task.pending_split_count:
        department_task.pending_split_count = max((department_task.pending_split_count or 0) - 1, 0)
        codes = list(department_task.pending_split_codes or [])
        department_task.pending_split_codes = codes[1:] if codes else []
    db.flush()
    sync_task_owners(db, task, owners)
    sync_sub_task_executors(db, task, executors)
    db.add(
        TaskEvent(
            object_type="sub_task",
            object_id=task.id,
            event_type="created",
            title="创建子任务",
            content=task.title,
            actor_id=current_user.id,
        )
    )
    db.commit()
    db.refresh(task)
    expire_task_people(db, task, include_executors=True)
    return serialize_sub_task(task)


@router.put("/sub-tasks/{sub_task_id}")
def update_sub_task(
    sub_task_id: int,
    payload: SubTaskUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> dict:
    task = db.get(SubTask, sub_task_id)
    if (
        not task
        or not task.department_task
        or task.department_task.status == "archived"
        or not task.department_task.parent_task
        or task.department_task.parent_task.status == "archived"
    ):
        raise HTTPException(status.HTTP_404_NOT_FOUND, detail="Sub task not found")
    if not can_split_sub_task(db, current_user, task.department_task):
        raise HTTPException(status.HTTP_403_FORBIDDEN, detail="No edit access to this sub task")
    data = payload.model_dump(exclude_unset=True)
    owner_ids = data.pop("owner_ids", None)
    owner_id = data.pop("owner_id", None)
    executor_ids = data.pop("executor_ids", None)
    executor_id = data.pop("executor_id", None)
    inherited_owners = owner_people(task.department_task)
    if not inherited_owners:
        raise HTTPException(status.HTTP_422_UNPROCESSABLE_ENTITY, detail="部门任务负责人不能为空")
    if owner_ids is not None or owner_id is not None:
        requested_payload = SubTaskUpdate(owner_ids=owner_ids, owner_id=owner_id)
        validate_inherited_sub_task_owners(requested_payload, inherited_owners)
    sync_task_owners(db, task, inherited_owners)
    if executor_ids is not None or executor_id is not None:
        sync_sub_task_executors(db, task, resolve_users(db, executor_ids, executor_id, "执行人"))
    for key, value in data.items():
        setattr(task, key, value)
    db.add(
        TaskEvent(
            object_type="sub_task",
            object_id=task.id,
            event_type="updated",
            title="编辑子任务",
            content=task.title,
            actor_id=current_user.id,
        )
    )
    db.commit()
    db.refresh(task)
    expire_task_people(db, task, include_executors=True)
    return serialize_sub_task(task)


@router.post("/sub-tasks/{sub_task_id}/start")
def start_sub_task(
    sub_task_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission("weekly_update.submit")),
) -> dict:
    task = db.get(SubTask, sub_task_id)
    if not task:
        raise HTTPException(status.HTTP_404_NOT_FOUND, detail="Sub task not found")
    if not can_update_sub_task_weekly(current_user, task):
        raise HTTPException(status.HTTP_403_FORBIDDEN, detail="No update access to this sub task")
    if task.status == "completed":
        raise HTTPException(status.HTTP_409_CONFLICT, detail="Completed task cannot be restarted")
    if task.status == "pending_update":
        task.status = "in_progress"
        task.started_at = datetime.now(timezone.utc)
        db.add(
            TaskEvent(
                object_type="sub_task",
                object_id=task.id,
                event_type="started",
                title="开启子任务",
                content=task.title,
                actor_id=current_user.id,
            )
        )
        db.commit()
        db.refresh(task)
    return serialize_sub_task(task)


@router.post("/sub-tasks/{sub_task_id}/complete")
def complete_sub_task(
    sub_task_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission("weekly_update.submit")),
) -> dict:
    task = db.get(SubTask, sub_task_id)
    if not task:
        raise HTTPException(status.HTTP_404_NOT_FOUND, detail="Sub task not found")
    if not can_update_sub_task_weekly(current_user, task):
        raise HTTPException(status.HTTP_403_FORBIDDEN, detail="No update access to this sub task")
    task.status = "completed"
    task.progress = 100
    db.add(
        TaskEvent(
            object_type="sub_task",
            object_id=task.id,
            event_type="completed",
            title="完成子任务",
            content=task.title,
            actor_id=current_user.id,
        )
    )
    db.commit()
    db.refresh(task)
    return serialize_sub_task(task)


@router.post("/sub-tasks/{sub_task_id}/reopen")
def reopen_sub_task(
    sub_task_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> dict:
    task = db.get(SubTask, sub_task_id)
    if not task or not task.department_task or task.department_task.status == "archived":
        raise HTTPException(status.HTTP_404_NOT_FOUND, detail="Sub task not found")
    if not can_reopen_sub_task(current_user, task):
        raise HTTPException(status.HTTP_403_FORBIDDEN, detail="No reopen access to this sub task")
    if task.status != "completed":
        raise HTTPException(status.HTTP_409_CONFLICT, detail="Only completed tasks can be reopened")
    task.status = "in_progress"
    task.progress = 0
    db.execute(
        update(WeeklyUpdate)
        .where(
            WeeklyUpdate.sub_task_id == task.id,
            WeeklyUpdate.week_key == current_week_key(),
        )
        .values(progress=0)
    )
    db.add(
        TaskEvent(
            object_type="sub_task",
            object_id=task.id,
            event_type="reopened",
            title="撤回子任务完成",
            content=task.title,
            actor_id=current_user.id,
        )
    )
    db.commit()
    db.refresh(task)
    return serialize_sub_task_for_execution(task, current_user)


@router.get("/weekly-updates")
def list_weekly_updates(
    week_key: str | None = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> list[dict]:
    stmt = select(WeeklyUpdate).order_by(WeeklyUpdate.id.desc())
    if week_key:
        stmt = stmt.where(WeeklyUpdate.week_key == week_key)
    updates = db.scalars(stmt).all()
    return [
        serialize_weekly_update_for_user(db, update, current_user)
        for update in updates
        if can_access_sub_task(db, current_user, update.sub_task)
    ]


@router.get("/weekly-updates/current")
def current_weekly_update(
    sub_task_id: int,
    assignee_id: int | None = None,
    week_key: str | None = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> dict:
    sub_task = db.get(SubTask, sub_task_id)
    if not sub_task or not sub_task.department_task or sub_task.department_task.status == "archived":
        raise HTTPException(status.HTTP_404_NOT_FOUND, detail="Sub task not found")
    if not can_access_sub_task(db, current_user, sub_task):
        raise HTTPException(status.HTTP_403_FORBIDDEN, detail="No access to this sub task")
    assignee = default_update_assignee(current_user, sub_task, assignee_id)
    target_week = week_key or current_week_key()
    update = db.scalar(
        select(WeeklyUpdate).where(
            WeeklyUpdate.sub_task_id == sub_task_id,
            WeeklyUpdate.week_key == target_week,
            WeeklyUpdate.assignee_id == assignee.id,
        )
    )
    if update:
        return serialize_weekly_update_for_user(db, update, current_user)
    return {
        "id": None,
        "sub_task_id": sub_task.id,
        "sub_task": sub_task.title,
        "assignee_id": assignee.id,
        "assignee": assignee.name,
        "week_key": target_week,
        "status": "empty",
        "progress": sub_task.progress or 0,
        "this_week": None,
        "next_week": None,
        "risk": None,
        "needs_coordination": False,
        "submitter_id": current_user.id,
        "submitter": current_user.name,
        "submitted_at": None,
        "attachments": [],
    }


@router.post("/weekly-updates")
def save_weekly_update(
    payload: WeeklyUpdateUpsert,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission("weekly_update.submit")),
) -> dict:
    sub_task = db.get(SubTask, payload.sub_task_id)
    if not sub_task or not sub_task.department_task or sub_task.department_task.status == "archived":
        raise HTTPException(status.HTTP_404_NOT_FOUND, detail="Sub task not found")
    assignee = default_update_assignee(current_user, sub_task, payload.assignee_id)
    if not can_update_sub_task_weekly(current_user, sub_task, assignee.id):
        raise HTTPException(status.HTTP_403_FORBIDDEN, detail="No update access to this sub task")
    if sub_task.status == "pending_update":
        raise HTTPException(status.HTTP_409_CONFLICT, detail="Sub task must be started before weekly update")
    if sub_task.status == "completed":
        raise HTTPException(status.HTTP_409_CONFLICT, detail="Completed task cannot be updated")
    update = upsert_weekly_update(
        db,
        sub_task=sub_task,
        user=current_user,
        assignee=assignee,
        week_key=payload.week_key,
        progress=payload.progress,
        this_week=payload.this_week,
        next_week=payload.next_week,
        risk=payload.risk,
        risk_level=None,
        needs_coordination=payload.needs_coordination,
        submit=payload.submit,
    )
    return serialize_weekly_update_for_user(db, update, current_user)


@router.post("/weekly-updates/{weekly_update_id}/attachments")
async def upload_weekly_update_attachment(
    weekly_update_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission("weekly_update.submit")),
) -> dict:
    update = db.get(WeeklyUpdate, weekly_update_id)
    if not update or not update.sub_task:
        raise HTTPException(status.HTTP_404_NOT_FOUND, detail="Weekly update not found")
    if not can_upload_weekly_update_attachment(current_user, update):
        raise HTTPException(status.HTTP_403_FORBIDDEN, detail="No upload access to this weekly update")
    content = await file.read(MAX_ATTACHMENT_BYTES + 1)
    if len(content) > MAX_ATTACHMENT_BYTES:
        raise HTTPException(status.HTTP_413_REQUEST_ENTITY_TOO_LARGE, detail="Attachment exceeds 20MB limit")
    filename = sanitize_filename(file.filename or "attachment")
    suffix = Path(filename).suffix[:20]
    if suffix.lower() not in ALLOWED_ATTACHMENT_EXTENSIONS:
        raise HTTPException(status.HTTP_422_UNPROCESSABLE_ENTITY, detail="Unsupported attachment file type")
    storage_dir = Path("weekly_updates") / str(update.id)
    stored_name = f"{uuid.uuid4().hex}{suffix}"
    root = attachment_storage_root()
    target_dir = (root / storage_dir).resolve()
    if not is_path_inside(target_dir, root):
        raise HTTPException(status.HTTP_400_BAD_REQUEST, detail="Invalid attachment path")
    target_dir.mkdir(parents=True, exist_ok=True)
    target_path = target_dir / stored_name
    target_path.write_bytes(content)
    attachment = Attachment(
        filename=filename,
        storage_path=str(storage_dir / stored_name),
        related_type="weekly_update",
        related_id=update.id,
        uploader_id=current_user.id,
    )
    db.add(attachment)
    db.commit()
    db.refresh(attachment)
    return serialize_attachment_for_user(attachment, current_user)


@router.get("/attachments/{attachment_id}/download")
def download_attachment(
    attachment_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> FileResponse:
    attachment = db.get(Attachment, attachment_id)
    if not attachment:
        raise HTTPException(status.HTTP_404_NOT_FOUND, detail="Attachment not found")
    if not can_access_attachment(db, current_user, attachment):
        raise HTTPException(status.HTTP_403_FORBIDDEN, detail="No download access to this attachment")
    root = attachment_storage_root()
    file_path = (root / attachment.storage_path).resolve()
    if not is_path_inside(file_path, root) or not file_path.exists() or not file_path.is_file():
        raise HTTPException(status.HTTP_404_NOT_FOUND, detail="Attachment file not found")
    return FileResponse(file_path, media_type="application/octet-stream", filename=attachment.filename)


@router.delete("/attachments/{attachment_id}")
def delete_attachment(
    attachment_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> dict:
    attachment = db.get(Attachment, attachment_id)
    if not attachment:
        raise HTTPException(status.HTTP_404_NOT_FOUND, detail="Attachment not found")
    if not can_delete_attachment(current_user, attachment):
        raise HTTPException(status.HTTP_403_FORBIDDEN, detail="No delete access to this attachment")
    root = attachment_storage_root()
    file_path = (root / attachment.storage_path).resolve()
    if is_path_inside(file_path, root) and file_path.exists() and file_path.is_file():
        file_path.unlink()
    db.delete(attachment)
    db.commit()
    return {"ok": True}


@router.get("/timeline")
def timeline(
    object_type: str | None = None,
    object_id: int | None = None,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    stmt = select(TaskEvent).order_by(TaskEvent.created_at.desc(), TaskEvent.id.desc())
    if object_type:
        stmt = stmt.where(TaskEvent.object_type == object_type)
    if object_id:
        stmt = stmt.where(TaskEvent.object_id == object_id)
    return [
        {
            "id": item.id,
            "object_type": item.object_type,
            "object_id": item.object_id,
            "event_type": item.event_type,
            "title": item.title,
            "content": item.content,
            "actor": item.actor.name if item.actor else None,
            "created_at": item.created_at.isoformat() if item.created_at else None,
        }
        for item in db.scalars(stmt).all()
    ]


@router.get("/timeline/matrix")
def timeline_matrix(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> dict:
    return build_timeline_matrix_payload(db, current_user)


@router.get("/risk-items")
def list_risk_items(
    status_filter: str | None = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> list[dict]:
    stmt = select(RiskItem).order_by(RiskItem.id.desc())
    if status_filter:
        stmt = stmt.where(RiskItem.status == status_filter)
    return [
        serialize_risk_item_for_user(item, current_user)
        for item in db.scalars(stmt).all()
        if item.sub_task and can_access_sub_task(db, current_user, item.sub_task)
    ]


@router.get("/risks")
def list_risks(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)) -> list[dict]:
    return list_risk_items(db=db, current_user=current_user)


@router.post("/risk-items", status_code=201)
async def create_risk_item(
    payload: RiskItemCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> dict:
    sub_task = db.get(SubTask, payload.sub_task_id)
    if not sub_task or not sub_task.department_task or sub_task.department_task.status == "archived":
        raise HTTPException(status.HTTP_404_NOT_FOUND, detail="Sub task not found")
    if not can_create_risk_item(current_user, sub_task):
        raise HTTPException(status.HTTP_403_FORBIDDEN, detail="No risk create access")
    source_update = db.get(WeeklyUpdate, payload.source_weekly_update_id) if payload.source_weekly_update_id else None
    if source_update and source_update.sub_task_id != sub_task.id:
        raise HTTPException(status.HTTP_422_UNPROCESSABLE_ENTITY, detail="风险来源周更新不属于该子任务")
    owners = risk_owner_candidates(sub_task)
    owner = owners[0] if owners else None
    if not owner:
        raise HTTPException(status.HTTP_422_UNPROCESSABLE_ENTITY, detail="子任务没有可用负责人，无法登记风险")
    score, level = risk_score(payload.impact_score, payload.likelihood_score)
    item = RiskItem(
        code=generate_code(db, RiskItem, "RI"),
        sub_task=sub_task,
        source_weekly_update=source_update,
        title=payload.title.strip(),
        description=payload.description,
        impact_score=payload.impact_score,
        likelihood_score=payload.likelihood_score,
        score=score,
        level=level,
        owner=owner,
        status="open",
        due_date=payload.due_date,
        created_by_id=current_user.id,
        updated_at=datetime.now(timezone.utc),
    )
    db.add(item)
    db.flush()
    db.add(
        TaskEvent(
            object_type="risk_item",
            object_id=item.id,
            event_type="created",
            title="新增风险项",
            content=f"{item.code} {item.title}",
            actor_id=current_user.id,
        )
    )
    db.commit()
    db.refresh(item)
    notification = None
    if item.level == "high":
        notification = await send_risk_item_notifications(db, item, "新增高风险")
    return {"risk_item": serialize_risk_item_for_user(item, current_user), "notification": notification}


@router.put("/risk-items/{risk_item_id}")
async def update_risk_item(
    risk_item_id: int,
    payload: RiskItemUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> dict:
    item = db.get(RiskItem, risk_item_id)
    if not item or not item.sub_task:
        raise HTTPException(status.HTTP_404_NOT_FOUND, detail="Risk item not found")
    if not can_manage_risk_item(current_user, item):
        raise HTTPException(status.HTTP_403_FORBIDDEN, detail="No risk manage access")
    old_level = item.level
    old_due_date = item.due_date
    data = payload.model_dump(exclude_unset=True)
    if "title" in data and data["title"]:
        item.title = data["title"].strip()
    if "description" in data:
        item.description = data["description"]
    if "impact_score" in data and data["impact_score"] is not None:
        item.impact_score = data["impact_score"]
    if "likelihood_score" in data and data["likelihood_score"] is not None:
        item.likelihood_score = data["likelihood_score"]
    item.score, item.level = risk_score(item.impact_score, item.likelihood_score)
    if "status" in data and data["status"]:
        if data["status"] not in {"open", "in_progress", "closed"}:
            raise HTTPException(status.HTTP_422_UNPROCESSABLE_ENTITY, detail="风险状态不合法")
        item.status = data["status"]
    if "owner_id" in data and data["owner_id"] is not None:
        owner_ids = {owner.id for owner in risk_owner_candidates(item.sub_task)}
        if data["owner_id"] not in owner_ids:
            raise HTTPException(status.HTTP_422_UNPROCESSABLE_ENTITY, detail="风险责任人必须是当前子任务负责人")
        item.owner_id = data["owner_id"]
    if "due_date" in data:
        item.due_date = data["due_date"]
    if "resolution_note" in data:
        item.resolution_note = data["resolution_note"]
    item.updated_at = datetime.now(timezone.utc)
    db.add(
        TaskEvent(
            object_type="risk_item",
            object_id=item.id,
            event_type="updated",
            title="更新风险项",
            content=f"{item.code} {item.status} {item.level}",
            actor_id=current_user.id,
        )
    )
    db.commit()
    db.refresh(item)
    notification = None
    if item.status in {"open", "in_progress"} and item.level == "high" and old_level != "high":
        notification = await send_risk_item_notifications(db, item, "升级为高风险")
    elif (
        item.status in {"open", "in_progress"}
        and item.due_date
        and item.due_date < date.today()
        and old_due_date != item.due_date
    ):
        notification = await send_risk_item_notifications(db, item, "风险逾期")
    return {"risk_item": serialize_risk_item_for_user(item, current_user), "notification": notification}


@router.get("/meeting-board")
def meeting_board(
    week_key: str | None = None,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
) -> dict:
    return build_meeting_board(db, week_key or current_week_key())


@router.get("/meeting-board/overview")
def meeting_board_overview(
    week_key: str | None = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> dict:
    return build_meeting_overview_payload(db, current_user, week_key or current_week_key())


@router.get("/meeting-board/parent")
def meeting_board_parent(
    week_key: str | None = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> dict:
    return build_parent_board_payload(db, current_user, week_key or current_week_key())


@router.get("/meeting-board/department")
def meeting_board_department(
    week_key: str | None = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> dict:
    return build_department_board_payload(db, current_user, week_key or current_week_key())


@router.get("/meeting-board/export")
def export_meeting_board(
    week_key: str | None = None,
    db: Session = Depends(get_db),
    _: User = Depends(require_permission("meeting.export")),
) -> Response:
    board = build_meeting_board(db, week_key or current_week_key())
    lines = [f"# {board['week_key']} 会议材料", ""]
    for section, title in [
        ("decision_items", "必须决策事项"),
        ("high_risks", "高风险任务"),
        ("missing_updates", "未更新任务"),
        ("completed", "本周完成事项"),
    ]:
        lines.append(f"## {title}")
        for item in board[section]:
            lines.append(f"- {item.get('title')}: {item.get('problem') or item.get('status', '')}")
        lines.append("")
    return Response("\n".join(lines), media_type="text/markdown; charset=utf-8")


@router.get("/notifications")
def list_notifications(
    include_historical: bool = False,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
) -> list[dict]:
    stmt = select(NotificationRecord)
    if not include_historical:
        stmt = stmt.where(
            NotificationRecord.notification_type.not_in(["lark_test_message", "weekly_update_reminder"]),
            or_(
                NotificationRecord.related_type.is_(None),
                NotificationRecord.related_type != "card_preview",
            ),
        )
    records = db.scalars(stmt.order_by(NotificationRecord.id.desc())).all()
    return [
        {
            "id": item.id,
            "target_user": item.target_user.name if item.target_user else None,
            "notification_type": item.notification_type,
            "related_type": item.related_type,
            "related_id": item.related_id,
            "title": item.title,
            "web_url": item.web_url,
            "send_status": item.send_status,
            "clicked": item.clicked,
            "first_clicked_at": item.first_clicked_at.isoformat() if item.first_clicked_at else None,
            "last_clicked_at": item.last_clicked_at.isoformat() if item.last_clicked_at else None,
            "click_count": item.click_count,
            "result": item.result,
            "dedupe_key": item.dedupe_key,
            "created_at": item.created_at.isoformat() if item.created_at else None,
        }
        for item in records
    ]


@router.get("/lark/diagnostics")
async def lark_diagnostics(_: User = Depends(require_permission("notification.nudge"))) -> dict:
    return await lark_client.health_check()


@router.post("/lark/resolve-open-ids")
async def resolve_lark_open_ids(
    db: Session = Depends(get_db),
    _: User = Depends(require_admin),
) -> dict:
    candidates = list(
        db.scalars(
            select(User)
            .where(User.email.is_not(None), User.open_id.is_(None), User.status != "disabled")
            .order_by(User.id)
        ).all()
    )
    results = []
    resolved = 0
    blocked = 0
    failed = 0
    for start in range(0, len(candidates), 50):
        batch = candidates[start : start + 50]
        emails = [user.email for user in batch if user.email]
        try:
            lookup = await lark_client.batch_get_user_ids_by_email(emails)
        except Exception as exc:
            failed += len(batch)
            for user in batch:
                results.append(
                    {
                        "user_id": user.id,
                        "name": user.name,
                        "email": user.email,
                        "status": "failed",
                        "message": str(exc),
                    }
                )
            continue

        users_by_email = lookup.get("users") or {}
        for user in batch:
            item = users_by_email.get(user.email or "")
            if not item:
                blocked += 1
                results.append(
                    {
                        "user_id": user.id,
                        "name": user.name,
                        "email": user.email,
                        "status": "blocked",
                        "message": "飞书未找到该邮箱",
                    }
                )
                continue
            remote_name = item.get("name") or item.get("user_name") or item.get("en_name")
            if remote_name and remote_name != user.name:
                blocked += 1
                results.append(
                    {
                        "user_id": user.id,
                        "name": user.name,
                        "email": user.email,
                        "status": "conflict",
                        "message": f"飞书姓名不一致：{remote_name}",
                    }
                )
                continue
            open_id = normalize_open_id(item.get("open_id") or item.get("user_id"))
            if not open_id:
                blocked += 1
                results.append(
                    {
                        "user_id": user.id,
                        "name": user.name,
                        "email": user.email,
                        "status": "blocked",
                        "message": "飞书返回结果缺少 open_id",
                    }
                )
                continue
            conflict = db.scalar(select(User).where(User.open_id == open_id, User.id != user.id))
            if conflict:
                blocked += 1
                results.append(
                    {
                        "user_id": user.id,
                        "name": user.name,
                        "email": user.email,
                        "status": "conflict",
                        "message": f"open_id 已绑定到 {conflict.name}",
                    }
                )
                continue
            user.open_id = open_id
            user.open_id_bound_at = datetime.now(timezone.utc)
            db.add(user)
            resolved += 1
            results.append(
                {
                    "user_id": user.id,
                    "name": user.name,
                    "email": user.email,
                    "status": "resolved",
                    "message": "已绑定 open_id",
                }
            )
    db.commit()
    return {
        "ok": failed == 0,
        "scanned": len(candidates),
        "resolved": resolved,
        "blocked": blocked,
        "failed": failed,
        "results": results,
    }


@router.post("/notifications/lark-weekly-reminders")
async def lark_weekly_reminders(
    payload: WeeklyReminderRequest,
    db: Session = Depends(get_db),
    _: User = Depends(require_permission("notification.nudge")),
) -> dict:
    return await send_weekly_update_reminders(db, payload.week_key)


@router.post("/notifications/department-task-due-reminders")
async def department_task_due_reminders(
    db: Session = Depends(get_db),
    _: User = Depends(require_permission("notification.nudge")),
) -> dict:
    return await send_department_task_due_reminders(db)


@router.get("/notifications/scheduler-status")
def notification_scheduler_status(
    _: User = Depends(require_permission("notification.nudge")),
) -> dict:
    return scheduler_status()


@router.post("/notifications/risk-overdue")
async def risk_overdue_notifications(
    db: Session = Depends(get_db),
    _: User = Depends(require_permission("notification.nudge")),
) -> dict:
    return await send_risk_overdue_reminders(db)


@router.get("/attachments")
def list_attachments(_: User = Depends(get_current_user)) -> list[dict]:
    return []


@router.get("/dashboard")
def dashboard(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)) -> dict:
    week_key = current_week_key()
    sub_tasks = list(db.scalars(select(SubTask)).all())
    updates = list(
        db.scalars(select(WeeklyUpdate).where(WeeklyUpdate.week_key == week_key)).all()
    )
    submitted = [item for item in updates if item.status == "submitted"]
    risk_items = active_risk_items_for_tasks(db, sub_tasks)
    overdue = [item for item in sub_tasks if item.due_date and item.due_date < __import__("datetime").date.today() and item.status != "completed"]
    return {
        "week_key": week_key,
        "current_user": serialize_user(current_user),
        "cards": {
            "parent_in_progress": db.scalar(
                select(__import__("sqlalchemy").func.count()).select_from(ParentTask).where(ParentTask.status == "in_progress")
            ),
            "weekly_due": len([item for item in sub_tasks if item.status != "completed"]),
            "risk_tasks": len(risk_items),
            "overdue_tasks": len(overdue),
        },
        "weekly_progress": {
            "expected": len([item for item in sub_tasks if item.status != "completed"]),
            "submitted": len(submitted),
            "missing": max(len([item for item in sub_tasks if item.status != "completed"]) - len(submitted), 0),
        },
        "risk_summary": {
            "high": len([item for item in risk_items if item.level == "high"]),
            "medium": len([item for item in risk_items if item.level == "medium"]),
            "low": len([item for item in risk_items if item.level == "low"]),
        },
    }


@router.get("/coordination-items")
def list_coordination_items(db: Session = Depends(get_db), _: User = Depends(get_current_user)) -> list[dict]:
    items = db.scalars(select(CoordinationItem).order_by(CoordinationItem.id.desc())).all()
    return [
        {
            "id": item.id,
            "title": item.title,
            "description": item.description,
            "status": item.status,
            "sub_task": item.sub_task.title if item.sub_task else None,
            "owner_id": item.owner_id,
        }
        for item in items
    ]
