import re
from collections import defaultdict
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.db.session import SessionLocal
from app.models.entities import (
    BaseSyncRun,
    Department,
    DepartmentTask,
    ParentTask,
    Role,
    StrategicGoal,
    SubTask,
    User,
    WeeklyUpdate,
)
from app.services.seed import clear_business_data

DEFAULT_EXCEL_PATH = Path("/data/base_download/2026公司工作任务跟踪表-3.xlsx")


def text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip()


def parse_date(value: Any) -> date | None:
    raw = text(value)
    match = re.search(r"(20\d{2})[-/](\d{1,2})[-/](\d{1,2})", raw)
    if not match:
        return None
    return date(int(match.group(1)), int(match.group(2)), int(match.group(3)))


def parse_week_key(value: Any) -> str:
    raw = text(value)
    match = re.search(r"W(\d{1,2})", raw, re.IGNORECASE)
    if match:
        return f"2026-W{int(match.group(1)):02d}"
    today = date.today()
    year, week, _ = today.isocalendar()
    return f"{year}-W{week:02d}"


def parse_progress(value: Any) -> int:
    raw = text(value)
    match = re.search(r"(\d{1,3})\s*%", raw)
    if match:
        return min(int(match.group(1)), 100)
    match = re.search(r"\d{1,3}", raw)
    return min(int(match.group(0)), 100) if match else 0


def parse_status(value: Any) -> str:
    raw = text(value)
    if "完成" in raw:
        return "completed"
    if "风险" in raw or "延期" in raw:
        return "risk"
    if "未启动" in raw:
        return "pending_update"
    return "in_progress"


def parse_risk_level(value: Any) -> str:
    raw = text(value)
    if any(word in raw for word in ["高", "延期", "风险", "阻塞"]):
        return "high"
    if "中" in raw:
        return "medium"
    if raw:
        return "low"
    return "none"


def strip_code_prefix(value: Any, code: str) -> str:
    raw = text(value)
    return re.sub(rf"^{re.escape(code)}\s*", "", raw).strip() or raw or code


def parent_code(value: Any) -> str:
    raw = text(value)
    match = re.match(r"(T-\d{3})\b", raw)
    return match.group(1) if match else raw


def split_people(value: Any) -> list[str]:
    raw = text(value)
    return [item.strip() for item in re.split(r"[\n,，、\s]+", raw) if item.strip()]


def split_departments(value: Any) -> list[str]:
    raw = text(value)
    return [item.strip() for item in re.split(r"[,，、\n]+", raw) if item.strip()]


def worksheet_rows(path: Path, sheet_name: str) -> list[dict[str, Any]]:
    workbook = load_workbook(path, data_only=True)
    sheet = workbook[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))
    headers = [text(value) for value in rows[0]]
    result = []
    for row in rows[1:]:
        if any(text(value) for value in row):
            result.append(dict(zip(headers, row)))
    workbook.close()
    return result


def ensure_department(db: Session, name: str) -> Department:
    clean = name or "未分配"
    department = db.scalar(select(Department).where(Department.name == clean))
    if not department:
        department = Department(name=clean, status="active")
        db.add(department)
        db.flush()
    return department


def ensure_user(db: Session, name: str, department: Department | None = None) -> User:
    clean = name or "待分配人员"
    user = db.scalar(select(User).where(User.name == clean))
    if not user:
        user = User(name=clean, department=department, status="pending", source="base_excel")
        db.add(user)
        db.flush()
    elif department and not user.department_id:
        user.department = department
    return user


def add_roles(db: Session, user: User, role_codes: list[str]) -> None:
    roles = db.scalars(select(Role).where(Role.code.in_(role_codes))).all()
    existing = {role.code for role in user.roles}
    for role in roles:
        if role.code not in existing:
            user.roles.append(role)


def import_excel_2026(db: Session, path: Path = DEFAULT_EXCEL_PATH, actor_id: int | None = None) -> dict:
    if not path.exists():
        raise FileNotFoundError(f"Excel export not found: {path}")

    goals_rows = worksheet_rows(path, "01_战略目标")
    department_rows = worksheet_rows(path, "02_部门任务")
    split_rows = worksheet_rows(path, "03_部门拆解任务")
    weekly_rows = worksheet_rows(path, "04_周更新进度")

    clear_business_data(db, include_sync_runs=False)

    goals: dict[str, StrategicGoal] = {}
    for row in goals_rows:
        code = text(row.get("目标编号"))
        goal = StrategicGoal(
            code=code,
            name=text(row.get("战略目标")),
            description=text(row.get("备注")),
            year=2026,
            progress=0,
            status=parse_status(row.get("状态")),
        )
        db.add(goal)
        db.flush()
        goals[code] = goal

    default_department = ensure_department(db, "未分配")
    default_owner = db.get(User, actor_id) if actor_id else db.scalar(select(User).where(User.is_admin.is_(True)))
    if not default_owner:
        default_owner = ensure_user(db, "系统导入人", default_department)

    rows_by_parent: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in department_rows:
        rows_by_parent[parent_code(row.get("母任务项"))].append(row)

    parent_tasks: dict[str, ParentTask] = {}
    for code, rows in rows_by_parent.items():
        first = rows[0]
        departments = split_departments(first.get("负责部门"))
        department = ensure_department(db, departments[0]) if departments else default_department
        owner_names = split_people(first.get("母任务负责人"))
        owners = [ensure_user(db, name, department) for name in owner_names] or [default_owner]
        owner = owners[0]
        for item in owners:
            add_roles(db, item, ["parent_owner"])
        due_dates = [parse_date(row.get("截止时间")) for row in rows]
        due_dates = [item for item in due_dates if item]
        task = ParentTask(
            code=code,
            title=strip_code_prefix(first.get("母任务项"), code),
            description=None,
            goal=goals.get(text(first.get("目标编号"))) or next(iter(goals.values())),
            department=department,
            owner=owner,
            priority="normal",
            status="in_progress",
            progress=0,
            due_date=max(due_dates) if due_dates else None,
        )
        task.owners = owners
        db.add(task)
        db.flush()
        parent_tasks[code] = task

    department_tasks: dict[str, DepartmentTask] = {}
    for row in department_rows:
        code = text(row.get("任务编号"))
        if not code:
            continue
        department_names = split_departments(row.get("负责部门"))
        departments = [ensure_department(db, name) for name in department_names] or [default_department]
        owner_names = split_people(row.get("任务负责人"))
        owners = [ensure_user(db, name, departments[0]) for name in owner_names] or [default_owner]
        owner = owners[0]
        for item in owners:
            add_roles(db, item, ["department_owner", "task_owner"])
        task = DepartmentTask(
            code=code,
            title=strip_code_prefix(row.get("任务项"), code),
            parent_task=parent_tasks[parent_code(row.get("母任务项"))],
            department=departments[0],
            departments=departments,
            owner=owner,
            owners=owners,
            status=parse_status(row.get("状态")),
            progress=0,
            due_date=parse_date(row.get("截止时间")),
            pending_split_count=0,
            pending_split_codes=[],
        )
        db.add(task)
        db.flush()
        department_tasks[code] = task

    sub_sequence: dict[str, int] = defaultdict(int)
    sub_tasks_by_weekly_key: dict[tuple[str, str], SubTask] = {}
    pending_by_task: dict[str, list[str]] = defaultdict(list)
    imported_weekly = 0
    skipped_weekly = 0

    for row in split_rows:
        task_code = text(row.get("任务编号"))
        original_sub_code = text(row.get("子任务编号"))
        concrete = text(row.get("具体任务"))
        department_task = department_tasks.get(task_code)
        if not department_task or not original_sub_code:
            continue
        if not concrete:
            pending_by_task[task_code].append(original_sub_code)
            continue
        sub_sequence[task_code] += 1
        code = f"{task_code}-{sub_sequence[task_code]:02d}"
        department = department_task.department
        owner_names = split_people(row.get("任务负责人（负责拆解任务到执行者）"))
        executor_names = split_people(row.get("执行责任人 (人员 )"))
        owners = [ensure_user(db, name, department) for name in owner_names] or list(department_task.owners or [department_task.owner])
        executors = [ensure_user(db, name, department) for name in executor_names] or [owners[0]]
        owner = owners[0]
        executor = executors[0]
        for item in owners:
            add_roles(db, item, ["task_owner"])
        for item in executors:
            add_roles(db, item, ["executor"])
        sub_task = SubTask(
            code=code,
            title=concrete,
            department_task=department_task,
            executor=executor,
            owner=owner,
            executors=executors,
            owners=owners,
            status=parse_status(row.get("状态")),
            progress=parse_progress(row.get("本周进度") or row.get("上周任务进度")),
            risk_level="none",
            due_date=parse_date(row.get("截止时间")),
        )
        db.add(sub_task)
        db.flush()
        sub_tasks_by_weekly_key[(text(row.get("任务项")), original_sub_code)] = sub_task
        sub_tasks_by_weekly_key[(task_code, original_sub_code)] = sub_task

    for task_code, codes in pending_by_task.items():
        task = department_tasks.get(task_code)
        if task:
            task.pending_split_count = len(codes)
            task.pending_split_codes = codes

    for row in weekly_rows:
        original_sub_code = text(row.get("子任务编号"))
        task_label = text(row.get("任务项"))
        sub_task = sub_tasks_by_weekly_key.get((task_label, original_sub_code))
        if not sub_task:
            task_code = task_label.split(" ", 1)[0]
            sub_task = sub_tasks_by_weekly_key.get((task_code, original_sub_code))
        if not sub_task:
            skipped_weekly += 1
            continue

        submitter = sub_task.executor
        risk_text = text(row.get("遗留事项"))
        update = WeeklyUpdate(
            sub_task=sub_task,
            week_key=parse_week_key(row.get("所属周")),
            status="submitted",
            progress=parse_progress(row.get("上周任务进度（自动更新）") or row.get("本周进度")),
            this_week=text(row.get("本周完成内容")),
            next_week=text(row.get("下周工作计划")),
            risk=risk_text,
            needs_coordination=False,
            submitter=submitter,
            submitted_at=datetime.now(timezone.utc),
        )
        db.add(update)
        sub_task.status = parse_status(row.get("状态"))
        sub_task.progress = max(sub_task.progress, update.progress)
        imported_weekly += 1

    for task in department_tasks.values():
        sub_tasks = task.sub_tasks
        if sub_tasks:
            task.progress = round(sum(item.progress for item in sub_tasks) / len(sub_tasks))
            if all(item.status == "completed" for item in sub_tasks):
                task.status = "completed"
            elif any(item.status == "risk" for item in sub_tasks):
                task.status = "risk"
            elif any(item.status == "in_progress" for item in sub_tasks):
                task.status = "in_progress"
            else:
                task.status = "pending_update"
        elif task.pending_split_count:
            task.status = "pending_split"

    for parent in parent_tasks.values():
        tasks = parent.department_tasks
        if tasks:
            parent.progress = round(sum(item.progress for item in tasks) / len(tasks))
            if all(item.status == "completed" for item in tasks):
                parent.status = "completed"
            elif any(item.status == "risk" for item in tasks):
                parent.status = "risk"
            else:
                parent.status = "in_progress"

    summary = {
        "goals": len(goals),
        "parent_tasks": len(parent_tasks),
        "department_tasks": len(department_tasks),
        "sub_tasks": sum(sub_sequence.values()),
        "pending_split_rows": sum(len(items) for items in pending_by_task.values()),
        "weekly_updates": imported_weekly,
        "skipped_weekly_updates": skipped_weekly,
        "risks": 0,
    }
    db.add(
        BaseSyncRun(
            source_name="2026公司工作任务跟踪表 Excel",
            table_name="01_战略目标,02_部门任务,03_部门拆解任务,04_周更新进度",
            status="success",
            record_count=summary["sub_tasks"] + summary["weekly_updates"],
            message="本地 Excel 重新导入，修正母任务/部门任务/子任务层级。",
            raw_summary=summary,
            actor_id=actor_id,
        )
    )
    db.commit()
    return summary


def main() -> None:
    import argparse
    import json

    parser = argparse.ArgumentParser()
    parser.add_argument("--path", default=str(DEFAULT_EXCEL_PATH))
    parser.add_argument("--actor-id", type=int, default=None)
    args = parser.parse_args()
    with SessionLocal() as db:
        summary = import_excel_2026(db, Path(args.path), args.actor_id)
    print(json.dumps({"ok": True, "summary": summary}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
