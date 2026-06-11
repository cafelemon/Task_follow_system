import argparse
import json
from collections import Counter, defaultdict
from io import BytesIO
from pathlib import Path
from typing import Any, BinaryIO

from openpyxl import load_workbook
from sqlalchemy import func, select, update
from sqlalchemy.orm import Session

from app.db.session import SessionLocal
from app.models.entities import (
    BaseSyncRun,
    Department,
    DepartmentTask,
    DepartmentTaskDepartment,
    ParentTask,
    Role,
    User,
)

REPO_ROOT = Path(__file__).resolve().parents[3]
CONTAINER_SOURCE_ROOT = Path("/data/base_download")
DEFAULT_CONTACT_FILENAME = "浙江势通机器人科技有限公司-通讯录-导出.xlsx"
COMPANY_CONTACT_NAMES = {"浙江势通机器人科技有限公司"}
COMPANY_CONTACT_EMAILS = {"shitong@citronbiotec.cn"}
EMAIL_COLUMN_NAMES = ("企业邮箱", "邮箱", "email", "Email", "工作邮箱")
NAME_COLUMN_NAMES = ("姓名", "name", "Name")
DEPARTMENT_RENAMES = {
    "研发部": "研发中心",
    "数据部": "数字与信息中心",
    "质量部": "质量体系部",
}


def default_contact_path() -> Path:
    container_path = CONTAINER_SOURCE_ROOT / DEFAULT_CONTACT_FILENAME
    return container_path if container_path.exists() else REPO_ROOT / DEFAULT_CONTACT_FILENAME


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_email(value: Any) -> str:
    return clean_text(value).lower()


def pick_value(row: dict[str, str], names: tuple[str, ...]) -> str:
    for name in names:
        if row.get(name):
            return row[name].strip()
    return ""


def read_contact_rows(source: Path | bytes | BinaryIO) -> list[dict[str, str]]:
    if isinstance(source, Path):
        workbook = load_workbook(source, read_only=True, data_only=True)
    elif isinstance(source, bytes):
        workbook = load_workbook(BytesIO(source), read_only=True, data_only=True)
    else:
        workbook = load_workbook(source, read_only=True, data_only=True)
    worksheet = workbook.active
    header_values = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
    header = [clean_text(value) for value in header_values]
    rows: list[dict[str, str]] = []
    for values in worksheet.iter_rows(min_row=2, values_only=True):
        row = {
            header[index]: clean_text(value)
            for index, value in enumerate(values)
            if index < len(header) and header[index]
        }
        if any(row.values()):
            rows.append(row)
    return rows


def empty_summary() -> dict[str, Any]:
    return {
        "version": "2.3.0",
        "source_rows": 0,
        "valid_contacts": 0,
        "skipped_company_contacts": 0,
        "email_updates": 0,
        "people_created": 0,
        "people_unchanged": 0,
        "blocked": 0,
        "open_id_candidates_after_apply": 0,
        "open_id_candidate_additions": 0,
        "department_migration": {
            old: {
                "target": new,
                "users": 0,
                "parent_tasks": 0,
                "department_tasks": 0,
                "department_task_department_updates": 0,
                "department_task_department_deletes": 0,
                "department_children": 0,
                "deleted": False,
            }
            for old, new in DEPARTMENT_RENAMES.items()
        },
        "results": [],
        "department_errors": [],
    }


def ensure_executor_role(db: Session) -> Role:
    role = db.scalar(select(Role).where(Role.code == "executor"))
    if not role:
        raise RuntimeError("缺少 executor 角色，无法为新增通讯录人员设置默认权限")
    return role


def count_where(db: Session, model: Any, *criteria: Any) -> int:
    return db.scalar(select(func.count()).select_from(model).where(*criteria)) or 0


def user_counts_by_name(db: Session) -> dict[str, list[User]]:
    grouped: dict[str, list[User]] = defaultdict(list)
    for user in db.scalars(select(User)).all():
        grouped[user.name].append(user)
    return grouped


def sync_people_from_rows(db: Session, rows: list[dict[str, str]], *, apply: bool) -> dict[str, Any]:
    summary = empty_summary()
    summary["source_rows"] = len(rows)
    names = [pick_value(row, NAME_COLUMN_NAMES) for row in rows]
    duplicate_source_names = {name for name, count in Counter(name for name in names if name).items() if count > 1}
    users_by_name = user_counts_by_name(db)
    executor_role = ensure_executor_role(db)

    for index, row in enumerate(rows, start=2):
        name = pick_value(row, NAME_COLUMN_NAMES)
        email = normalize_email(pick_value(row, EMAIL_COLUMN_NAMES))
        base = {"row": index, "name": name, "email": email}

        if name in COMPANY_CONTACT_NAMES or email in COMPANY_CONTACT_EMAILS:
            summary["skipped_company_contacts"] += 1
            summary["results"].append({**base, "status": "skipped", "message": "公司主体邮箱已跳过"})
            continue
        if not name or not email:
            summary["blocked"] += 1
            summary["results"].append({**base, "status": "blocked", "message": "姓名或邮箱为空"})
            continue
        summary["valid_contacts"] += 1
        if name in duplicate_source_names:
            summary["blocked"] += 1
            summary["results"].append({**base, "status": "conflict", "message": "导入文件内姓名重复"})
            continue

        existing_email_user = db.scalar(select(User).where(User.email == email))
        matches = users_by_name.get(name, [])
        if len(matches) > 1:
            summary["blocked"] += 1
            summary["results"].append({**base, "status": "conflict", "message": "系统内姓名重复"})
            continue

        if matches:
            user = matches[0]
            if existing_email_user and existing_email_user.id != user.id:
                summary["blocked"] += 1
                summary["results"].append(
                    {**base, "user_id": user.id, "status": "conflict", "message": f"邮箱已绑定到 {existing_email_user.name}"}
                )
                continue
            if user.email == email:
                summary["people_unchanged"] += 1
                summary["results"].append({**base, "user_id": user.id, "status": "unchanged", "message": "邮箱已存在"})
                continue
            if user.email and user.email != email:
                summary["blocked"] += 1
                summary["results"].append(
                    {**base, "user_id": user.id, "status": "conflict", "message": f"系统内已有不同邮箱：{user.email}"}
                )
                continue
            summary["email_updates"] += 1
            if user.open_id is None and user.status != "disabled":
                summary["open_id_candidate_additions"] += 1
            summary["results"].append({**base, "user_id": user.id, "status": "updated", "message": "已补齐邮箱"})
            if apply:
                user.email = email
                db.add(user)
            continue

        if existing_email_user:
            summary["blocked"] += 1
            summary["results"].append({**base, "status": "conflict", "message": f"邮箱已绑定到 {existing_email_user.name}"})
            continue

        summary["people_created"] += 1
        summary["open_id_candidate_additions"] += 1
        summary["results"].append({**base, "status": "created", "message": "新增 pending 子任务执行者"})
        if apply:
            user = User(
                name=name,
                email=email,
                department_id=None,
                status="pending",
                source="contact_import_2_3_0",
            )
            user.roles = [executor_role]
            db.add(user)
            db.flush()
            users_by_name[name] = [user]

    return summary


def migrate_department_task_departments(
    db: Session,
    old_department: Department,
    target_department: Department,
    *,
    apply: bool,
) -> tuple[int, int]:
    rows = list(
        db.scalars(
            select(DepartmentTaskDepartment).where(DepartmentTaskDepartment.department_id == old_department.id)
        ).all()
    )
    updates = 0
    deletes = 0
    for row in rows:
        target_exists = db.get(
            DepartmentTaskDepartment,
            {"department_task_id": row.department_task_id, "department_id": target_department.id},
        )
        if target_exists:
            deletes += 1
            if apply:
                db.delete(row)
        else:
            updates += 1
            if apply:
                row.department_id = target_department.id
                db.add(row)
    return updates, deletes


def migrate_departments(db: Session, summary: dict[str, Any], *, apply: bool) -> None:
    for old_name, target_name in DEPARTMENT_RENAMES.items():
        item = summary["department_migration"][old_name]
        old_department = db.scalar(select(Department).where(Department.name == old_name))
        target_department = db.scalar(select(Department).where(Department.name == target_name))
        if not old_department:
            item["deleted"] = True
            continue
        if not target_department:
            summary["department_errors"].append(f"缺少目标部门：{target_name}")
            continue

        item["users"] = count_where(db, User, User.department_id == old_department.id)
        item["parent_tasks"] = count_where(db, ParentTask, ParentTask.department_id == old_department.id)
        item["department_tasks"] = count_where(db, DepartmentTask, DepartmentTask.department_id == old_department.id)
        item["department_children"] = count_where(db, Department, Department.parent_id == old_department.id)
        updates, deletes = migrate_department_task_departments(db, old_department, target_department, apply=apply)
        item["department_task_department_updates"] = updates
        item["department_task_department_deletes"] = deletes

        if apply:
            db.execute(update(User).where(User.department_id == old_department.id).values(department_id=target_department.id))
            db.execute(update(ParentTask).where(ParentTask.department_id == old_department.id).values(department_id=target_department.id))
            db.execute(update(DepartmentTask).where(DepartmentTask.department_id == old_department.id).values(department_id=target_department.id))
            db.execute(update(Department).where(Department.parent_id == old_department.id).values(parent_id=target_department.id))
            db.flush()
            residuals = {
                "users": count_where(db, User, User.department_id == old_department.id),
                "parent_tasks": count_where(db, ParentTask, ParentTask.department_id == old_department.id),
                "department_tasks": count_where(db, DepartmentTask, DepartmentTask.department_id == old_department.id),
                "department_task_departments": count_where(
                    db, DepartmentTaskDepartment, DepartmentTaskDepartment.department_id == old_department.id
                ),
                "department_children": count_where(db, Department, Department.parent_id == old_department.id),
            }
            if any(residuals.values()):
                summary["department_errors"].append(f"{old_name} 迁移后仍有引用：{residuals}")
                continue
            db.delete(old_department)
            item["deleted"] = True


def count_open_id_candidates(db: Session) -> int:
    return count_where(db, User, User.email.is_not(None), User.open_id.is_(None), User.status != "disabled")


def sync_people_and_departments(
    db: Session,
    *,
    rows: list[dict[str, str]],
    apply: bool,
    actor_id: int | None = None,
    source_name: str = DEFAULT_CONTACT_FILENAME,
) -> dict[str, Any]:
    summary = sync_people_from_rows(db, rows, apply=apply)
    migrate_departments(db, summary, apply=apply)
    if summary["department_errors"]:
        raise RuntimeError("; ".join(summary["department_errors"]))

    if apply:
        db.flush()
        summary["open_id_candidates_after_apply"] = count_open_id_candidates(db)
        db.add(
            BaseSyncRun(
                source_name=source_name,
                table_name="通讯录",
                status="success",
                record_count=summary["valid_contacts"],
                message="2.3.0 通讯录人员邮箱补齐、旧部门迁移清理。",
                raw_summary=summary,
                actor_id=actor_id,
            )
        )
        db.commit()
    else:
        summary["open_id_candidates_after_apply"] = count_open_id_candidates(db) + summary["open_id_candidate_additions"]
        db.rollback()
    return summary


def run_from_path(path: Path, *, apply: bool, actor_id: int | None = None) -> dict[str, Any]:
    rows = read_contact_rows(path)
    with SessionLocal() as db:
        return sync_people_and_departments(
            db,
            rows=rows,
            apply=apply,
            actor_id=actor_id,
            source_name=path.name,
        )


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--contacts", type=Path, default=default_contact_path())
    mode = parser.add_mutually_exclusive_group(required=True)
    mode.add_argument("--dry-run", action="store_true")
    mode.add_argument("--apply", action="store_true")
    args = parser.parse_args()
    summary = run_from_path(args.contacts, apply=args.apply)
    print(json.dumps(summary, ensure_ascii=False, indent=2, sort_keys=True))


if __name__ == "__main__":
    main()
